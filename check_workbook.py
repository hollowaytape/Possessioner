from openpyxl import load_workbook

wb = load_workbook("D:\\Code\\roms\\Possessioner\\PSSR_dump.xlsx")
ws = wb["YUMI.MSD"]

# Find the record at offset 0x01b2d
for row_num in range(2, ws.max_row + 1):
    offset_cell = ws.cell(row_num, 1).value  # Offset column
    if offset_cell and "01b2d" in str(offset_cell).lower():
        print(f"Found record at row {row_num}:")
        for col_num in range(1, 15):
            header = ws.cell(1, col_num).value
            value = ws.cell(row_num, col_num).value
            print(f"  {header}: {value}")
        print()
