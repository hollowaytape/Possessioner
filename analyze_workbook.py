from openpyxl import load_workbook
import json

wb = load_workbook("D:\\Code\\roms\\Possessioner\\PSSR_dump.xlsx")

# Check for POS1.MSD and YUMI.MSD sheets
for sheet_name in ["POS1.MSD", "YUMI.MSD"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        print(f"Max row: {ws.max_row}")
        
        # Get column headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"Columns: {headers}\n")
        
        # Find Command column index
        command_idx = None
        offset_idx = None
        for i, header in enumerate(headers):
            if header and "Command" in str(header):
                command_idx = i + 1  # openpyxl is 1-indexed
            if header and "Offset" in str(header):
                offset_idx = i + 1
        
        if command_idx:
            print(f"Command column index: {command_idx}")
            for row_num in range(2, min(50, ws.max_row + 1)):
                offset_cell = ws.cell(row_num, offset_idx).value if offset_idx else None
                command_cell = ws.cell(row_num, command_idx).value
                if command_cell:
                    print(f"  Offset {offset_cell}: {command_cell}")
