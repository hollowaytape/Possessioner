from openpyxl import load_workbook
from rominfo import MSD_POINTER_RANGES

def get_command_from_pointer(msd_file, pointer_loc):
    """Get command name from pointer location in workbook."""
    wb = load_workbook("D:\\Code\\roms\\Possessioner\\PSSR_pointer_dump.xlsx")
    
    if msd_file not in wb.sheetnames:
        return None
    
    ws = wb[msd_file]
    
    # Columns: A=Text Offset, B=Pointer, C=Text
    for row in range(2, ws.max_row + 1):
        pointer_cell = ws.cell(row, 2).value
        text_offset_cell = ws.cell(row, 1).value
        
        if pointer_cell is None:
            continue
        
        try:
            ptr_val = int(str(pointer_cell).replace("0x", ""), 16) if "0x" in str(pointer_cell) else None
            if ptr_val and ptr_val == pointer_loc:
                return text_offset_cell
        except:
            pass
    
    return None

# Try to get data from PSSR_dump workbook for command names
wb_dump = load_workbook("D:\\Code\\roms\\Possessioner\\PSSR_dump.xlsx")

flag_details = {}

for msd_file in ["POS1.MSD", "YUMI.MSD"]:
    print(f"\n=== {msd_file} FLAG ANALYSIS ===\n")
    
    if msd_file not in wb_dump.sheetnames:
        continue
    
    ws = wb_dump[msd_file]
    
    # Create a mapping from offset to command
    offset_to_command = {}
    for row in range(2, ws.max_row + 1):
        offset_cell = ws.cell(row, 1).value
        command_cell = ws.cell(row, 2).value
        if offset_cell and command_cell:
            try:
                offset_val = int(str(offset_cell).replace("0x", ""), 16) if "0x" in str(offset_cell) else None
                if offset_val is not None:
                    offset_to_command[offset_val] = command_cell
            except:
                pass
    
    # Now match flag operations to nearby commands
    ranges = MSD_POINTER_RANGES.get(msd_file, [])
    for start, end in ranges:
        print(f"Range 0x{start:05x} - 0x{end:05x}:")
        
        # Find all commands in this range
        cmds_in_range = sorted([(off, cmd) for off, cmd in offset_to_command.items() if start <= off <= end])
        
        for offset, cmd in cmds_in_range[:5]:  # Show first 5 commands
            print(f"  Command at 0x{offset:05x}: {cmd}")

