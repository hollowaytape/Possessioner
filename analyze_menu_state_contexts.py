from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rominfo import DUMP_XLS_PATH, MSD_POINTER_RANGES, ORIGINAL_ROM_DIR

MENU_STATE_PATTERN = re.compile(
    bytes.fromhex("01 05 00 11 05 11 06 11 07 00 11")
    + b"."
    + bytes.fromhex("00 01 11")
    + b"."
    + bytes.fromhex("00 01 01 07 00"),
    re.DOTALL,
)


def parse_offset(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    return int(text, 16)


def hex5(value: int | None) -> str | None:
    if value is None:
        return None
    return f"0x{value:05x}"


def load_dump_rows(path: Path) -> dict[tuple[str, int], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows_by_file_offset: dict[tuple[str, int], dict[str, Any]] = {}

    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith(".MSD"):
            continue
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(rows)]
        except StopIteration:
            continue
        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        if "Offset" not in name_to_index:
            continue

        current_command = ""
        for row in rows:
            offset = parse_offset(row[name_to_index["Offset"]])
            if offset is None:
                continue
            command = row[name_to_index["Command"]] if "Command" in name_to_index else None
            if command:
                current_command = str(command).strip()
            rows_by_file_offset[(sheet_name, offset)] = {
                "file": sheet_name,
                "offset": offset,
                "command": str(command).strip() if command else "",
                "block_command": current_command,
                "english": row[name_to_index["English"]] if "English" in name_to_index else None,
            }

    workbook.close()
    return rows_by_file_offset


def build_offset_lookup(rows_by_file_offset: dict[tuple[str, int], dict[str, Any]]) -> dict[int, list[dict[str, Any]]]:
    result: dict[int, list[dict[str, Any]]] = {}
    for (_, offset), row in rows_by_file_offset.items():
        result.setdefault(offset, []).append(row)
    return result


def owner_file(location: int) -> str | None:
    for filename, ranges in MSD_POINTER_RANGES.items():
        for start, stop in ranges:
            if start <= location <= stop:
                return filename
    return None


def parse_text_entries(data: bytes, start: int, stop: int) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    cursor = start
    while cursor + 4 < stop and cursor + 4 < len(data):
        if data[cursor] != 0x02 or data[cursor + 1] == 0xFF:
            break
        text_offset = data[cursor + 1] | (data[cursor + 2] << 8)
        line_count = data[cursor + 3]
        terminator = data[cursor + 4]
        entries.append(
            {
                "opcode_location": cursor,
                "text_offset": text_offset,
                "line_count": line_count,
                "terminator": terminator,
            }
        )
        cursor += 5
        if terminator != 0xFF:
            break
    return entries


def parse_flag_ops(data: bytes, start: int, stop: int) -> list[dict[str, Any]]:
    operations: list[dict[str, Any]] = []
    cursor = start
    while cursor + 4 < stop and cursor + 4 < len(data):
        kind = None
        op_length = 0
        if data[cursor : cursor + 3] == b"\x02\xff\x01":
            kind = "set_flag"
            op_length = 5
        elif data[cursor : cursor + 3] == b"\x02\xff\x02":
            kind = "check_flag"
            op_length = 5
        elif data[cursor : cursor + 3] == b"\x03\xff\x01" and cursor + 5 < len(data):
            kind = "clear_flag"
            op_length = 6

        if kind is None:
            break

        op = {
            "location": cursor,
            "kind": kind,
            "arg1": data[cursor + 3],
            "arg2": data[cursor + 4],
        }
        if op_length == 6:
            op["extra"] = data[cursor + 5]
        operations.append(op)
        cursor += op_length
    return operations


def scan_menu_state_blocks(pos_exe: bytes, offset_lookup: dict[int, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for match in MENU_STATE_PATTERN.finditer(pos_exe):
        location = match.start()
        slot = pos_exe[location + 11]
        text_entries = parse_text_entries(pos_exe, location + match.end() - match.start(), min(location + 0x60, len(pos_exe)))
        flag_ops = parse_flag_ops(
            pos_exe,
            text_entries[-1]["opcode_location"] + 5 if text_entries else location + match.end() - match.start(),
            min(location + 0x80, len(pos_exe)),
        )
        resolved_entries = []
        for entry in text_entries:
            rows = offset_lookup.get(entry["text_offset"], [])
            resolved_entries.append(
                {
                    "opcode_location": hex5(entry["opcode_location"]),
                    "text_offset": hex5(entry["text_offset"]),
                    "line_count": entry["line_count"],
                    "terminator": f"0x{entry['terminator']:02x}",
                    "rows": [
                        {
                            "file": row["file"],
                            "offset": hex5(row["offset"]),
                            "command": row["command"],
                            "block_command": row["block_command"],
                            "english": row["english"],
                        }
                        for row in rows
                    ],
                }
            )

        blocks.append(
            {
                "location": hex5(location),
                "owner_file": owner_file(location),
                "slot": slot,
                "slot_hex": f"0x{slot:02x}",
                "pattern_bytes": pos_exe[location : location + 21].hex(" "),
                "text_entries": resolved_entries,
                "flag_ops": [
                    {
                        "location": hex5(op["location"]),
                        "kind": op["kind"],
                        "arg1": f"0x{op['arg1']:02x}",
                        "arg2": f"0x{op['arg2']:02x}",
                        **({"extra": f"0x{op['extra']:02x}"} if "extra" in op else {}),
                    }
                    for op in flag_ops
                ],
            }
        )
    return blocks


def human_report(blocks: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    lines.append(f"Menu-state block count: {len(blocks)}")
    lines.append("")
    for block in blocks:
        lines.append(f"{block['location']} owner={block['owner_file']} slot={block['slot_hex']}")
        lines.append(f"  Pattern: {block['pattern_bytes']}")
        lines.append("  Text entries:")
        for entry in block["text_entries"]:
            if entry["rows"]:
                labels = ", ".join(
                    f"{row['file']} {row['offset']} -> {row['block_command'] or row['command'] or '(no command)'}"
                    for row in entry["rows"][:3]
                )
            else:
                labels = "(no workbook rows)"
            lines.append(
                f"    {entry['opcode_location']} -> {entry['text_offset']} lines={entry['line_count']} term={entry['terminator']} :: {labels}"
            )
        lines.append("  Flag ops:")
        if block["flag_ops"]:
            for op in block["flag_ops"]:
                extra = f" extra={op['extra']}" if "extra" in op else ""
                lines.append(f"    {op['location']} {op['kind']} {op['arg1']}:{op['arg2']}{extra}")
        else:
            lines.append("    (none)")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze POS.EXE menu-state blocks like the POS1 local talk-state machine.")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of human-readable output")
    args = parser.parse_args()

    dump_rows = load_dump_rows(Path(DUMP_XLS_PATH))
    offset_lookup = build_offset_lookup(dump_rows)
    pos_exe = (Path(ORIGINAL_ROM_DIR) / "POS.EXE").read_bytes()
    blocks = scan_menu_state_blocks(pos_exe, offset_lookup)

    if args.json:
        print(json.dumps(blocks, indent=2, ensure_ascii=False))
    else:
        print(human_report(blocks))


if __name__ == "__main__":
    main()
