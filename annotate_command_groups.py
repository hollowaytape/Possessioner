from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from analyze_dispatch_contexts import build_text_context_lookup, classify_command, scan_dispatch_headers
from analyze_handoff_contexts import scan_handoff_context_lookup, summarize_handoff_details
from command_groups import (
    find_arrival_groups,
    find_command_groups,
    find_direct_range_groups,
    format_offset,
    load_pos_exe_bytes,
    parse_offset,
)
from rominfo import DUMP_XLS_PATH, POINTER_DUMP_XLS_PATH

GROUP_LABEL_HEADER = "Command Group Label"
GROUP_START_HEADER = "Command Group Start"
GROUP_SIZE_HEADER = "Command Group Size"
DISPLAY_TYPE_HEADER = "Command Display Type"
DISPLAY_DETAIL_HEADER = "Command Display Detail"
DISPATCH_ACTION_HEADER = "Command Dispatch Action"
DISPATCH_TARGET_HEADER = "Command Dispatch Target"
DISPATCH_HEADER_HEADER = "Command Dispatch Header"
DISPATCH_DETAIL_HEADER = "Command Dispatch Detail"
HANDOFF_TYPE_HEADER = "Command Handoff Type"
HANDOFF_DESTINATION_HEADER = "Command Handoff Destination"
HANDOFF_DETAIL_HEADER = "Command Handoff Detail"
HEADER_WIDTHS = {
    GROUP_LABEL_HEADER: 30,
    GROUP_START_HEADER: 12,
    GROUP_SIZE_HEADER: 8,
    DISPLAY_TYPE_HEADER: 22,
    DISPLAY_DETAIL_HEADER: 50,
    DISPATCH_ACTION_HEADER: 18,
    DISPATCH_TARGET_HEADER: 18,
    DISPATCH_HEADER_HEADER: 16,
    DISPATCH_DETAIL_HEADER: 60,
    HANDOFF_TYPE_HEADER: 22,
    HANDOFF_DESTINATION_HEADER: 32,
    HANDOFF_DETAIL_HEADER: 60,
}


def resolve_header_columns(headers: list[object]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    next_column = len(headers) + 1
    for header_name in (
        GROUP_LABEL_HEADER,
        GROUP_START_HEADER,
        GROUP_SIZE_HEADER,
        DISPLAY_TYPE_HEADER,
        DISPLAY_DETAIL_HEADER,
        DISPATCH_ACTION_HEADER,
        DISPATCH_TARGET_HEADER,
        DISPATCH_HEADER_HEADER,
        DISPATCH_DETAIL_HEADER,
        HANDOFF_TYPE_HEADER,
        HANDOFF_DESTINATION_HEADER,
        HANDOFF_DETAIL_HEADER,
    ):
        if header_name in headers:
            column_map[header_name] = headers.index(header_name) + 1
        else:
            column_map[header_name] = next_column
            next_column += 1
    return column_map


def load_pointer_rows(path: Path) -> dict[str, dict[int, list[int]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    result: dict[str, dict[int, list[int]]] = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(rows)]
        except StopIteration:
            continue

        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        if "Text Loc" not in name_to_index or "Ptr Loc" not in name_to_index:
            continue

        sheet_rows: dict[int, list[int]] = {}
        for row in rows:
            text_loc = parse_offset(row[name_to_index["Text Loc"]])
            ptr_loc = parse_offset(row[name_to_index["Ptr Loc"]])
            if text_loc is None or ptr_loc is None:
                continue
            sheet_rows.setdefault(text_loc, []).append(ptr_loc)
        result[sheet_name] = sheet_rows

    workbook.close()
    return result


def find_direct_ranges_from_pointer_rows(
    row_offsets: list[int],
    pointer_rows: dict[int, list[int]],
    pos_exe_bytes: bytes,
) -> dict[int, list[tuple[int, int, int]]]:
    grouped_rows: dict[int, dict[tuple[int, int, int], None]] = {}
    offset_to_index = {offset: index for index, offset in enumerate(row_offsets)}

    for text_offset, pointer_locations in pointer_rows.items():
        start_index = offset_to_index.get(text_offset)
        if start_index is None:
            continue
        for pointer_location in pointer_locations:
            if pointer_location < 1 or pointer_location + 5 >= len(pos_exe_bytes):
                continue
            if pos_exe_bytes[pointer_location - 1] != 0xBE:
                continue
            if pos_exe_bytes[pointer_location + 2] != 0xB9:
                continue
            if pos_exe_bytes[pointer_location + 5] != 0x9A:
                continue
            if pos_exe_bytes[pointer_location + 6 : pointer_location + 10] != b"\x46\x22\xf0\x05":
                continue

            group_size = pos_exe_bytes[pointer_location + 3] | (pos_exe_bytes[pointer_location + 4] << 8)
            if group_size <= 0:
                continue

            group_rows = row_offsets[start_index : start_index + group_size]
            if len(group_rows) != group_size:
                continue

            for row_offset in group_rows:
                grouped_rows.setdefault(row_offset, {})[(text_offset, group_size, pointer_location - 1)] = None

    return {
        row_offset: sorted(group_map.keys(), key=lambda item: (item[0], item[2]))
        for row_offset, group_map in grouped_rows.items()
    }


def summarize_dispatch_details(dispatch_contexts: list[dict[str, object]]) -> list[str]:
    grouped: dict[tuple[object, ...], dict[str, object]] = {}

    for context in dispatch_contexts:
        action = str(context.get("action") or "").strip()
        target = str(context.get("target") or "").strip()
        command_label = str(context.get("command") or "").strip()
        header_location = context.get("header_location")
        key = (header_location, action, target, command_label)
        entry = grouped.setdefault(
            key,
            {
                "inline": False,
                "via_locations": [],
            },
        )
        via_target_location = context.get("via_target_location")
        if isinstance(via_target_location, int):
            if via_target_location not in entry["via_locations"]:
                entry["via_locations"].append(via_target_location)
        else:
            entry["inline"] = True

    details: list[str] = []
    for (header_location, action, target, command_label), entry in grouped.items():
        route_bits = []
        if action:
            route_bits.append(action)
        if target:
            route_bits.append(target)
        route_text = " / ".join(route_bits) if route_bits else "dispatch context"
        if command_label:
            route_text += f" -> {command_label}"

        qualifier_bits = []
        if isinstance(header_location, int):
            qualifier_bits.append(f"header {format_offset(header_location)}")
        if entry["inline"]:
            qualifier_bits.append("inline")
        if entry["via_locations"]:
            qualifier_bits.append(
                "via " + ", ".join(format_offset(location) for location in entry["via_locations"])
            )

        if qualifier_bits:
            route_text += " [" + "; ".join(qualifier_bits) + "]"
        details.append(route_text)

    return details


def apply_header_format_openpyxl(worksheet, header_columns: dict[str, int]) -> None:
    template_cell = worksheet.cell(row=1, column=1)
    default_font = Font(bold=True)
    default_fill = PatternFill(fill_type="solid", fgColor="808080")
    default_alignment = Alignment(horizontal="center", vertical="center")
    default_border = Border(bottom=copy(template_cell.border.bottom))

    for header_name, column in header_columns.items():
        cell = worksheet.cell(row=1, column=column)
        cell.font = copy(template_cell.font) if template_cell.has_style else copy(default_font)
        cell.fill = copy(template_cell.fill) if template_cell.has_style else copy(default_fill)
        cell.alignment = copy(template_cell.alignment) if template_cell.has_style else copy(default_alignment)
        cell.border = copy(template_cell.border) if template_cell.has_style else copy(default_border)
        width = HEADER_WIDTHS.get(header_name)
        if width is not None:
            worksheet.column_dimensions[get_column_letter(column)].width = width


def apply_header_format_excel(worksheet, header_columns: dict[str, int]) -> None:
    for header_name, column in header_columns.items():
        cell = worksheet.Cells(1, column)
        cell.Font.Bold = True
        cell.HorizontalAlignment = -4108  # xlCenter
        cell.VerticalAlignment = -4108  # xlCenter
        cell.Interior.Color = 0x808080
        width = HEADER_WIDTHS.get(header_name)
        if width is not None:
            worksheet.Columns(column).ColumnWidth = width


def collect_sheet_updates(
    worksheet,
    pos_exe_bytes: bytes,
    pointer_rows_by_sheet: dict[str, dict[int, list[int]]],
) -> tuple[dict[str, int], dict[int, tuple[str, str, str, str, str, str, str, str, str]]]:
    headers = [cell.value for cell in worksheet[1]]
    if "Offset" not in headers:
        return {}, {}
    if "Command" not in headers:
        return {}, {}

    offset_index = headers.index("Offset") + 1
    command_index = headers.index("Command") + 1
    header_columns = resolve_header_columns(headers)

    row_offsets: list[int] = []
    row_numbers_by_offset: dict[int, int] = {}
    command_by_offset: dict[int, str] = {}
    block_command_by_offset: dict[int, str] = {}
    current_block_command = ""

    for row_number in range(2, worksheet.max_row + 1):
        offset_value = worksheet.cell(row=row_number, column=offset_index).value
        offset = parse_offset(offset_value)
        if offset is None:
            continue
        row_offsets.append(offset)
        row_numbers_by_offset[offset] = row_number
        command_value = worksheet.cell(row=row_number, column=command_index).value
        command_text = str(command_value).strip() if command_value else ""
        command_by_offset[offset] = command_text
        if command_text:
            current_block_command = command_text
        block_command_by_offset[offset] = current_block_command

    _, grouped_rows = find_command_groups(worksheet.title, row_offsets, pos_exe_bytes)
    arrival_rows = find_arrival_groups(worksheet.title, row_offsets, pos_exe_bytes)
    direct_range_rows = find_direct_range_groups(worksheet.title, row_offsets, pos_exe_bytes)
    direct_range_rows_from_pointers = find_direct_ranges_from_pointer_rows(
        row_offsets,
        pointer_rows_by_sheet.get(worksheet.title, {}),
        pos_exe_bytes,
    )
    manual_block_by_offset: dict[int, tuple[int, str]] = {}
    current_manual_block: tuple[int, str] | None = None
    for offset in row_offsets:
        command_name = command_by_offset.get(offset, "")
        if command_name:
            current_manual_block = (offset, command_name)
        if current_manual_block is not None:
            manual_block_by_offset[offset] = current_manual_block

    dispatch_headers = scan_dispatch_headers(
        worksheet.title,
        pos_exe_bytes,
        [
            {
                "offset": offset,
                "command": command_by_offset.get(offset, ""),
                "block_command": block_command_by_offset.get(offset, ""),
                "english": None,
            }
            for offset in row_offsets
        ],
    )
    dispatch_context_lookup = build_text_context_lookup(dispatch_headers)
    dispatch_action_target_lookup: dict[tuple[str, str], list[dict[str, object]]] = {}
    for header in dispatch_headers:
        for target_context in header["candidate_targets"]:
            action = str(target_context.get("inferred_action") or "").strip()
            target = str(target_context.get("inferred_target") or "").strip()
            if not action and not target:
                continue
            dispatch_action_target_lookup.setdefault((action, target), []).append(
                {
                    "header_location": header["header_location"],
                    "via_target_location": target_context["target_location"],
                    "action": action,
                    "target": target,
                    "command": "",
                }
            )

    handoff_context_lookup = scan_handoff_context_lookup(
        worksheet.title,
        pos_exe_bytes,
        [
            {
                "file": worksheet.title,
                "offset": offset,
                "command": command_by_offset.get(offset, ""),
                "block_command": block_command_by_offset.get(offset, ""),
                "english": None,
            }
            for offset in row_offsets
        ],
    )

    updates: dict[int, tuple[str, str, str, str, str, str, str, str, str, str, str, str]] = {}
    for offset in row_offsets:
        row_number = row_numbers_by_offset[offset]
        groups = grouped_rows.get(offset, [])
        if groups:
            labels = []
            starts = []
            sizes = []
            for start_offset, size in groups:
                start_text = format_offset(start_offset)
                starts.append(start_text)
                sizes.append(str(size))
                command_name = command_by_offset.get(start_offset, "")
                if command_name:
                    labels.append(f"{command_name} @ {start_text}")
                else:
                    labels.append(start_text)
            label_value = "; ".join(labels)
            start_value = "; ".join(starts)
            size_value = "; ".join(sizes)
        else:
            label_value = ""
            start_value = ""
            size_value = ""

        display_types: list[str] = []
        display_details: list[str] = []

        if groups:
            for start_offset, size in groups:
                display_types.append("0x02 text command")
                display_details.append(f"0x02 group {format_offset(start_offset)} size {size}")

        for start_offset, size, pointer_location in arrival_rows.get(offset, []):
            display_types.append("arrival text")
            display_details.append(
                f"arrival {format_offset(start_offset)} size {size} via {format_offset(pointer_location)}"
            )

        for start_offset, size, call_location in direct_range_rows.get(offset, []):
            display_types.append("direct range call")
            display_details.append(
                f"direct range {format_offset(start_offset)} size {size} via {format_offset(call_location)}"
            )
        for start_offset, size, call_location in direct_range_rows_from_pointers.get(offset, []):
            detail = f"direct range {format_offset(start_offset)} size {size} via {format_offset(call_location)}"
            if detail not in display_details:
                display_types.append("direct range call")
                display_details.append(detail)

        if not display_types and offset == 0 and command_by_offset.get(offset) == "(Arrive)":
            display_types.append("arrival text")
            display_details.append("arrival offset 0 (no pointer override known)")

        if not display_types:
            manual_block = manual_block_by_offset.get(offset)
            if manual_block is not None:
                block_start, block_command = manual_block
                if block_command == "?":
                    display_types.append("no known display path")
                    display_details.append(
                        f"manual ? block {format_offset(block_start)} (no 0x02/arrival/direct-range trigger found)"
                    )
                elif block_command:
                    display_types.append("manual block only")
                    display_details.append(
                        f"manual block {block_command} @ {format_offset(block_start)} (no 0x02/arrival/direct-range trigger found)"
                    )

        display_type_value = "; ".join(display_types)
        display_detail_value = "; ".join(display_details)

        dispatch_contexts: list[dict[str, object]] = []
        seen_dispatch_keys: set[tuple[object, ...]] = set()
        dispatch_start_offsets = {offset}
        dispatch_start_offsets.update(start_offset for start_offset, _size in groups)
        dispatch_start_offsets.update(start_offset for start_offset, _size, _pointer_location in arrival_rows.get(offset, []))
        dispatch_start_offsets.update(start_offset for start_offset, _size, _call_location in direct_range_rows.get(offset, []))
        dispatch_start_offsets.update(
            start_offset for start_offset, _size, _call_location in direct_range_rows_from_pointers.get(offset, [])
        )

        for start_offset in sorted(dispatch_start_offsets):
            for context in dispatch_context_lookup.get(start_offset, []):
                key = (
                    context["header_location"],
                    context["action"],
                    context["target"],
                    context["command"],
                    context["via_target_location"] is None,
                )
                if key in seen_dispatch_keys:
                    continue
                seen_dispatch_keys.add(key)
                dispatch_contexts.append(context)

        if not dispatch_contexts:
            fallback_action, fallback_target = classify_command(block_command_by_offset.get(offset, ""))
            for context in dispatch_action_target_lookup.get((fallback_action, fallback_target), []):
                key = (
                    context["header_location"],
                    context["action"],
                    context["target"],
                    context["command"],
                    context["via_target_location"] is None,
                )
                if key in seen_dispatch_keys:
                    continue
                seen_dispatch_keys.add(key)
                dispatch_contexts.append(context)

        dispatch_actions: list[str] = []
        dispatch_targets: list[str] = []
        dispatch_headers: list[str] = []
        dispatch_details: list[str] = []
        for context in dispatch_contexts:
            action = str(context.get("action") or "").strip()
            target = str(context.get("target") or "").strip()
            header_location = context.get("header_location")

            if action and action not in dispatch_actions:
                dispatch_actions.append(action)
            if target and target not in dispatch_targets:
                dispatch_targets.append(target)

            header_text = format_offset(header_location) if isinstance(header_location, int) else ""
            if header_text and header_text not in dispatch_headers:
                dispatch_headers.append(header_text)

        dispatch_details = summarize_dispatch_details(dispatch_contexts)

        dispatch_action_value = "; ".join(dispatch_actions)
        dispatch_target_value = "; ".join(dispatch_targets)
        dispatch_header_value = "; ".join(dispatch_headers)
        dispatch_detail_value = "; ".join(dispatch_details)

        handoff_contexts: list[dict[str, object]] = []
        seen_handoff_keys: set[tuple[object, ...]] = set()
        handoff_start_offsets = {offset}
        handoff_start_offsets.update(start_offset for start_offset, _size in groups)
        handoff_start_offsets.update(start_offset for start_offset, _size, _pointer_location in arrival_rows.get(offset, []))
        handoff_start_offsets.update(start_offset for start_offset, _size, _call_location in direct_range_rows.get(offset, []))
        handoff_start_offsets.update(
            start_offset for start_offset, _size, _call_location in direct_range_rows_from_pointers.get(offset, [])
        )

        for start_offset in sorted(handoff_start_offsets):
            for context in handoff_context_lookup.get(start_offset, []):
                key = (context["location"], context["arg1"], context["destination_name"], context["handoff_type"])
                if key in seen_handoff_keys:
                    continue
                seen_handoff_keys.add(key)
                handoff_contexts.append(context)

        handoff_types: list[str] = []
        handoff_destinations: list[str] = []
        for context in handoff_contexts:
            handoff_type = str(context.get("handoff_type") or "").strip()
            destination_name = str(context.get("destination_label") or "").strip()
            if handoff_type and handoff_type not in handoff_types:
                handoff_types.append(handoff_type)
            if destination_name and destination_name not in handoff_destinations:
                handoff_destinations.append(destination_name)

        handoff_type_value = "; ".join(handoff_types)
        handoff_destination_value = "; ".join(handoff_destinations)
        handoff_detail_value = "; ".join(summarize_handoff_details(handoff_contexts))

        updates[row_number] = (
            label_value,
            start_value,
            size_value,
            display_type_value,
            display_detail_value,
            dispatch_action_value,
            dispatch_target_value,
            dispatch_header_value,
            dispatch_detail_value,
            handoff_type_value,
            handoff_destination_value,
            handoff_detail_value,
        )

    return {
        GROUP_LABEL_HEADER: header_columns[GROUP_LABEL_HEADER],
        GROUP_START_HEADER: header_columns[GROUP_START_HEADER],
        GROUP_SIZE_HEADER: header_columns[GROUP_SIZE_HEADER],
        DISPLAY_TYPE_HEADER: header_columns[DISPLAY_TYPE_HEADER],
        DISPLAY_DETAIL_HEADER: header_columns[DISPLAY_DETAIL_HEADER],
        DISPATCH_ACTION_HEADER: header_columns[DISPATCH_ACTION_HEADER],
        DISPATCH_TARGET_HEADER: header_columns[DISPATCH_TARGET_HEADER],
        DISPATCH_HEADER_HEADER: header_columns[DISPATCH_HEADER_HEADER],
        DISPATCH_DETAIL_HEADER: header_columns[DISPATCH_DETAIL_HEADER],
        HANDOFF_TYPE_HEADER: header_columns[HANDOFF_TYPE_HEADER],
        HANDOFF_DESTINATION_HEADER: header_columns[HANDOFF_DESTINATION_HEADER],
        HANDOFF_DETAIL_HEADER: header_columns[HANDOFF_DETAIL_HEADER],
    }, updates


def apply_updates_openpyxl(workbook, pos_exe_bytes: bytes, pointer_rows_by_sheet: dict[str, dict[int, list[int]]]) -> int:
    total_rows = 0
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if not sheet_name.endswith(".MSD"):
            continue
        header_columns, updates = collect_sheet_updates(worksheet, pos_exe_bytes, pointer_rows_by_sheet)
        if not header_columns:
            continue
        worksheet.cell(row=1, column=header_columns[GROUP_LABEL_HEADER], value=GROUP_LABEL_HEADER)
        worksheet.cell(row=1, column=header_columns[GROUP_START_HEADER], value=GROUP_START_HEADER)
        worksheet.cell(row=1, column=header_columns[GROUP_SIZE_HEADER], value=GROUP_SIZE_HEADER)
        worksheet.cell(row=1, column=header_columns[DISPLAY_TYPE_HEADER], value=DISPLAY_TYPE_HEADER)
        worksheet.cell(row=1, column=header_columns[DISPLAY_DETAIL_HEADER], value=DISPLAY_DETAIL_HEADER)
        worksheet.cell(row=1, column=header_columns[DISPATCH_ACTION_HEADER], value=DISPATCH_ACTION_HEADER)
        worksheet.cell(row=1, column=header_columns[DISPATCH_TARGET_HEADER], value=DISPATCH_TARGET_HEADER)
        worksheet.cell(row=1, column=header_columns[DISPATCH_HEADER_HEADER], value=DISPATCH_HEADER_HEADER)
        worksheet.cell(row=1, column=header_columns[DISPATCH_DETAIL_HEADER], value=DISPATCH_DETAIL_HEADER)
        worksheet.cell(row=1, column=header_columns[HANDOFF_TYPE_HEADER], value=HANDOFF_TYPE_HEADER)
        worksheet.cell(row=1, column=header_columns[HANDOFF_DESTINATION_HEADER], value=HANDOFF_DESTINATION_HEADER)
        worksheet.cell(row=1, column=header_columns[HANDOFF_DETAIL_HEADER], value=HANDOFF_DETAIL_HEADER)
        apply_header_format_openpyxl(worksheet, header_columns)
        for row_number, (
            label_value,
            start_value,
            size_value,
            display_type_value,
            display_detail_value,
            dispatch_action_value,
            dispatch_target_value,
            dispatch_header_value,
            dispatch_detail_value,
            handoff_type_value,
            handoff_destination_value,
            handoff_detail_value,
        ) in updates.items():
            worksheet.cell(row=row_number, column=header_columns[GROUP_LABEL_HEADER], value=label_value)
            worksheet.cell(row=row_number, column=header_columns[GROUP_START_HEADER], value=start_value)
            worksheet.cell(row=row_number, column=header_columns[GROUP_SIZE_HEADER], value=size_value)
            worksheet.cell(row=row_number, column=header_columns[DISPLAY_TYPE_HEADER], value=display_type_value)
            worksheet.cell(row=row_number, column=header_columns[DISPLAY_DETAIL_HEADER], value=display_detail_value)
            worksheet.cell(row=row_number, column=header_columns[DISPATCH_ACTION_HEADER], value=dispatch_action_value)
            worksheet.cell(row=row_number, column=header_columns[DISPATCH_TARGET_HEADER], value=dispatch_target_value)
            worksheet.cell(row=row_number, column=header_columns[DISPATCH_HEADER_HEADER], value=dispatch_header_value)
            worksheet.cell(row=row_number, column=header_columns[DISPATCH_DETAIL_HEADER], value=dispatch_detail_value)
            worksheet.cell(row=row_number, column=header_columns[HANDOFF_TYPE_HEADER], value=handoff_type_value)
            worksheet.cell(row=row_number, column=header_columns[HANDOFF_DESTINATION_HEADER], value=handoff_destination_value)
            worksheet.cell(row=row_number, column=header_columns[HANDOFF_DETAIL_HEADER], value=handoff_detail_value)
        total_rows += len(updates)
        print(f"{sheet_name}: updated {len(updates)} rows")
    return total_rows


def apply_updates_excel(
    xlsx_path: Path,
    pos_exe_bytes: bytes,
    pointer_rows_by_sheet: dict[str, dict[int, list[int]]],
) -> int:
    import win32com.client

    workbook_ro = load_workbook(xlsx_path, read_only=True, data_only=False)
    sheet_updates: dict[str, tuple[dict[str, int], dict[int, tuple[str, str, str, str, str, str, str, str, str, str, str, str]]]] = {}
    for sheet_name in workbook_ro.sheetnames:
        worksheet = workbook_ro[sheet_name]
        if not sheet_name.endswith(".MSD"):
            continue
        header_columns, updates = collect_sheet_updates(worksheet, pos_exe_bytes, pointer_rows_by_sheet)
        if header_columns:
            sheet_updates[sheet_name] = (header_columns, updates)
            print(f"{sheet_name}: updated {len(updates)} rows")
    workbook_ro.close()

    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        excel = win32com.client.Dispatch("Excel.Application")
    workbook = None
    full_name = str(xlsx_path.resolve()).lower()
    for candidate in excel.Workbooks:
        if str(candidate.FullName).lower() == full_name:
            workbook = candidate
            break
    if workbook is None:
        workbook = excel.Workbooks.Open(str(xlsx_path.resolve()))

    total_rows = 0
    for sheet_name, (header_columns, updates) in sheet_updates.items():
        worksheet = workbook.Worksheets(sheet_name)
        worksheet.Cells(1, header_columns[GROUP_LABEL_HEADER]).Value = GROUP_LABEL_HEADER
        worksheet.Cells(1, header_columns[GROUP_START_HEADER]).Value = GROUP_START_HEADER
        worksheet.Cells(1, header_columns[GROUP_SIZE_HEADER]).Value = GROUP_SIZE_HEADER
        worksheet.Cells(1, header_columns[DISPLAY_TYPE_HEADER]).Value = DISPLAY_TYPE_HEADER
        worksheet.Cells(1, header_columns[DISPLAY_DETAIL_HEADER]).Value = DISPLAY_DETAIL_HEADER
        worksheet.Cells(1, header_columns[DISPATCH_ACTION_HEADER]).Value = DISPATCH_ACTION_HEADER
        worksheet.Cells(1, header_columns[DISPATCH_TARGET_HEADER]).Value = DISPATCH_TARGET_HEADER
        worksheet.Cells(1, header_columns[DISPATCH_HEADER_HEADER]).Value = DISPATCH_HEADER_HEADER
        worksheet.Cells(1, header_columns[DISPATCH_DETAIL_HEADER]).Value = DISPATCH_DETAIL_HEADER
        worksheet.Cells(1, header_columns[HANDOFF_TYPE_HEADER]).Value = HANDOFF_TYPE_HEADER
        worksheet.Cells(1, header_columns[HANDOFF_DESTINATION_HEADER]).Value = HANDOFF_DESTINATION_HEADER
        worksheet.Cells(1, header_columns[HANDOFF_DETAIL_HEADER]).Value = HANDOFF_DETAIL_HEADER
        apply_header_format_excel(worksheet, header_columns)
        for row_number, (
            label_value,
            start_value,
            size_value,
            display_type_value,
            display_detail_value,
            dispatch_action_value,
            dispatch_target_value,
            dispatch_header_value,
            dispatch_detail_value,
            handoff_type_value,
            handoff_destination_value,
            handoff_detail_value,
        ) in updates.items():
            worksheet.Cells(row_number, header_columns[GROUP_LABEL_HEADER]).Value = label_value
            worksheet.Cells(row_number, header_columns[GROUP_START_HEADER]).Value = start_value
            worksheet.Cells(row_number, header_columns[GROUP_SIZE_HEADER]).Value = size_value
            worksheet.Cells(row_number, header_columns[DISPLAY_TYPE_HEADER]).Value = display_type_value
            worksheet.Cells(row_number, header_columns[DISPLAY_DETAIL_HEADER]).Value = display_detail_value
            worksheet.Cells(row_number, header_columns[DISPATCH_ACTION_HEADER]).Value = dispatch_action_value
            worksheet.Cells(row_number, header_columns[DISPATCH_TARGET_HEADER]).Value = dispatch_target_value
            worksheet.Cells(row_number, header_columns[DISPATCH_HEADER_HEADER]).Value = dispatch_header_value
            worksheet.Cells(row_number, header_columns[DISPATCH_DETAIL_HEADER]).Value = dispatch_detail_value
            worksheet.Cells(row_number, header_columns[HANDOFF_TYPE_HEADER]).Value = handoff_type_value
            worksheet.Cells(row_number, header_columns[HANDOFF_DESTINATION_HEADER]).Value = handoff_destination_value
            worksheet.Cells(row_number, header_columns[HANDOFF_DETAIL_HEADER]).Value = handoff_detail_value
        total_rows += len(updates)

    workbook.Save()
    return total_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Annotate PSSR_dump.xlsx with command-group metadata.")
    parser.add_argument("--xlsx", type=Path, default=Path(DUMP_XLS_PATH), help=f"Workbook to annotate (default: {DUMP_XLS_PATH})")
    args = parser.parse_args()

    pos_exe_bytes = load_pos_exe_bytes()
    pointer_rows_by_sheet = load_pointer_rows(Path(POINTER_DUMP_XLS_PATH))
    workbook = load_workbook(args.xlsx)

    total_rows = apply_updates_openpyxl(workbook, pos_exe_bytes, pointer_rows_by_sheet)
    try:
        workbook.save(args.xlsx)
    except PermissionError:
        workbook.close()
        print("Workbook is locked; falling back to live Excel automation")
        total_rows = apply_updates_excel(args.xlsx, pos_exe_bytes, pointer_rows_by_sheet)
    else:
        workbook.close()
    print(f"Updated {total_rows} rows total")


if __name__ == "__main__":
    main()
