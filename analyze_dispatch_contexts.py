from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rominfo import DUMP_XLS_PATH, MSD_POINTER_RANGES, ORIGINAL_ROM_DIR


DISPLAY_CALL = b"\x9a\x46\x22\xf0\x05"
HEADER = b"\xff\xff\xff\x01"
TARGET_SCAN_BYTES = 0x24
SUBHANDLER_SCAN_BYTES = 0x80
ACTION_PREFIXES = ("Look", "Talk", "Examine", "Think")


def parse_offset(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    return int(text, 16)


def hex5(value: int | None) -> str | None:
    if value is None:
        return None
    return f"0x{value:05x}"


def load_dump_rows(path: Path, file_filter: set[str]) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name not in file_filter:
            continue

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(rows)]
        except StopIteration:
            continue

        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        if "Offset" not in name_to_index:
            continue

        current_command = ""
        sheet_rows: list[dict[str, Any]] = []
        for row in rows:
            offset = parse_offset(row[name_to_index["Offset"]])
            if offset is None:
                continue

            command_value = row[name_to_index["Command"]] if "Command" in name_to_index else None
            if command_value:
                current_command = str(command_value).strip()

            sheet_rows.append(
                {
                    "offset": offset,
                    "command": str(command_value).strip() if command_value else "",
                    "block_command": current_command,
                    "english": row[name_to_index["English"]] if "English" in name_to_index else None,
                }
            )

        result[sheet_name] = sheet_rows

    workbook.close()
    return result


def build_offset_lookup(sheet_rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    return {row["offset"]: row for row in sheet_rows}


def find_next_display(data: bytes, start: int, stop: int) -> dict[str, Any] | None:
    cursor = max(0, start)
    limit = min(stop, len(data) - 1)
    while cursor <= limit:
        if cursor + 3 < len(data) and data[cursor] == 0x02 and data[cursor + 1] != 0xFF:
            text_offset = data[cursor + 1] | (data[cursor + 2] << 8)
            return {
                "kind": "text",
                "opcode_location": cursor,
                "text_offset": text_offset,
                "arg": data[cursor + 3],
            }

        if (
            cursor + 10 < len(data)
            and data[cursor] == 0xBE
            and data[cursor + 3] == 0xB9
            and data[cursor + 6 : cursor + 11] == DISPLAY_CALL
        ):
            text_offset = data[cursor + 1] | (data[cursor + 2] << 8)
            row_count = data[cursor + 4] | (data[cursor + 5] << 8)
            return {
                "kind": "direct_range",
                "opcode_location": cursor,
                "text_offset": text_offset,
                "row_count": row_count,
            }
        cursor += 1
    return None


def find_displays(data: bytes, start: int, stop: int, limit: int = 4) -> list[dict[str, Any]]:
    displays: list[dict[str, Any]] = []
    cursor = max(0, start)
    end = min(stop, len(data) - 1)
    while cursor <= end and len(displays) < limit:
        display = find_next_display(data, cursor, end)
        if display is None:
            break
        displays.append(display)
        next_cursor = display["opcode_location"] + (4 if display["kind"] == "text" else 11)
        if next_cursor <= cursor:
            next_cursor = cursor + 1
        cursor = next_cursor
    return displays


def find_flag_ops(data: bytes, start: int, stop: int, limit: int = 6) -> list[dict[str, Any]]:
    cursor = max(0, start)
    end = min(stop, len(data) - 6)
    operations: list[dict[str, Any]] = []
    while cursor <= end and len(operations) < limit:
        op = None
        length = 0

        if data[cursor : cursor + 3] == b"\x02\xff\x01":
            op = {
                "location": cursor,
                "kind": "set_flag",
                "arg1": data[cursor + 3],
                "arg2": data[cursor + 4],
            }
            length = 5
        elif data[cursor : cursor + 3] == b"\x02\xff\x02":
            op = {
                "location": cursor,
                "kind": "check_flag",
                "arg1": data[cursor + 3],
                "arg2": data[cursor + 4],
            }
            length = 5
        elif data[cursor : cursor + 3] == b"\x03\xff\x01":
            op = {
                "location": cursor,
                "kind": "clear_flag",
                "arg1": data[cursor + 3],
                "arg2": data[cursor + 4],
                "extra": data[cursor + 5],
            }
            length = 6

        if op is not None:
            operations.append(op)
            cursor += length
        else:
            cursor += 1

    return operations


def is_inside_flag_opcode(data: bytes, location: int) -> bool:
    for back in (1, 2, 3, 4):
        start = location - back
        if start < 0:
            continue
        if data[start : start + 3] == b"\x02\xff\x01" and start < location < start + 5:
            return True
        if data[start : start + 3] == b"\x02\xff\x02" and start < location < start + 5:
            return True
        if data[start : start + 3] == b"\x03\xff\x01" and start < location < start + 6:
            return True
    return False


def normalize_target_name(command: str) -> str:
    cleaned = re.sub(r"\s*\([^)]*\)", "", command).strip(" -")
    if not cleaned:
        return ""
    words = cleaned.split()
    if not words:
        return ""
    if len(words) >= 2 and words[0] in ACTION_PREFIXES and words[1] == "-":
        words = words[2:]
    elif words[0] in ACTION_PREFIXES:
        words = words[1:]
    if not words:
        return ""
    return words[0]


def classify_command(command: str) -> tuple[str, str]:
    if not command:
        return "", ""
    command = command.strip()
    if command == "(Arrive)":
        return "Arrive", "Arrive"
    for prefix in ACTION_PREFIXES:
        if command.startswith(prefix):
            return prefix, normalize_target_name(command)
    return "", ""


def is_plausible_script_target(data: bytes, target: int, stop: int) -> bool:
    if target < 0 or target >= len(data) or target >= stop:
        return False
    if data[target] in {0x00, 0x01, 0x02, 0x03, 0x04, 0x05, 0x06, 0x07, 0x08, 0xFF}:
        return True
    return find_next_display(data, target, min(target + 32, stop)) is not None


def find_candidate_relative_targets(data: bytes, header_location: int, stop: int) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    seen_targets: set[int] = set()
    table_end = min(header_location + 4 + TARGET_SCAN_BYTES, stop - 1)

    for cursor in range(header_location + 4, table_end, 2):
        if cursor + 1 >= len(data):
            break
        relative_offset = data[cursor] | (data[cursor + 1] << 8)
        if relative_offset == 0:
            continue
        target_location = header_location + relative_offset
        if target_location in seen_targets:
            continue
        if not is_plausible_script_target(data, target_location, stop):
            continue
        seen_targets.add(target_location)
        candidates.append(
            {
                "table_location": cursor,
                "relative_offset": relative_offset,
                "target_location": target_location,
            }
        )

    return candidates


def summarize_displays(
    displays: list[dict[str, Any]],
    offset_lookup: dict[int, dict[str, Any]],
    pos_exe: bytes,
) -> tuple[list[dict[str, Any]], Counter[str], Counter[str]]:
    summaries: list[dict[str, Any]] = []
    action_counter: Counter[str] = Counter()
    target_counter: Counter[str] = Counter()

    for display in displays:
        if display["kind"] == "text" and is_inside_flag_opcode(pos_exe, display["opcode_location"]):
            continue
        row = offset_lookup.get(display["text_offset"])
        if row is None:
            continue
        command = row["block_command"] if row else ""
        action, target = classify_command(command)
        if action:
            action_counter[action] += 1
        if target:
            target_counter[target] += 1
        summaries.append(
            {
                "kind": display["kind"],
                "opcode_location": display["opcode_location"],
                "text_offset": display["text_offset"],
                "arg": display.get("arg"),
                "row_count": display.get("row_count"),
                "command": row["command"] if row else "",
                "block_command": command,
                "english": row["english"] if row else None,
                "action": action,
                "target": target,
            }
        )

    return summaries, action_counter, target_counter


def scan_dispatch_headers(
    filename: str,
    pos_exe: bytes,
    sheet_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    offset_lookup = build_offset_lookup(sheet_rows)
    headers: list[dict[str, Any]] = []

    for range_start, range_stop in MSD_POINTER_RANGES[filename]:
        range_headers = [
            location
            for location in range(range_start, min(range_stop, len(pos_exe) - 4) + 1)
            if pos_exe[location : location + 4] == HEADER
        ]

        for index, header_location in enumerate(range_headers):
            next_header = range_headers[index + 1] if index + 1 < len(range_headers) else None
            scan_stop = min(range_stop + 1, header_location + 0x200, len(pos_exe))
            candidate_targets = find_candidate_relative_targets(pos_exe, header_location, scan_stop)

            inline_displays_raw = find_displays(
                pos_exe,
                header_location + 4,
                min(header_location + 0x60, scan_stop),
                limit=6,
            )
            inline_displays, inline_actions, inline_targets = summarize_displays(
                inline_displays_raw,
                offset_lookup,
                pos_exe,
            )

            subhandlers: list[dict[str, Any]] = []
            action_counter: Counter[str] = Counter(inline_actions)
            target_counter: Counter[str] = Counter(inline_targets)
            text_contexts: list[dict[str, Any]] = []

            for target in candidate_targets:
                displays_raw = find_displays(
                    pos_exe,
                    target["target_location"],
                    min(target["target_location"] + SUBHANDLER_SCAN_BYTES, scan_stop),
                    limit=4,
                )
                display_summaries, display_actions, display_targets = summarize_displays(
                    displays_raw,
                    offset_lookup,
                    pos_exe,
                )
                action_counter.update(display_actions)
                target_counter.update(display_targets)

                flags = find_flag_ops(
                    pos_exe,
                    target["target_location"],
                    min(target["target_location"] + 0x30, scan_stop),
                    limit=6,
                )

                inferred_action = display_actions.most_common(1)[0][0] if display_actions else ""
                inferred_target = display_targets.most_common(1)[0][0] if display_targets else ""

                subhandlers.append(
                    {
                        "table_location": target["table_location"],
                        "relative_offset": target["relative_offset"],
                        "target_location": target["target_location"],
                        "flags": flags,
                        "displays": display_summaries,
                        "inferred_action": inferred_action,
                        "inferred_target": inferred_target,
                    }
                )

                for display in display_summaries:
                    text_contexts.append(
                        {
                            "text_offset": display["text_offset"],
                            "opcode_location": display["opcode_location"],
                            "header_location": header_location,
                            "via_target_location": target["target_location"],
                            "action": inferred_action or display["action"],
                            "target": inferred_target or display["target"],
                            "command": display["block_command"],
                        }
                    )

            for display in inline_displays:
                text_contexts.append(
                    {
                        "text_offset": display["text_offset"],
                        "opcode_location": display["opcode_location"],
                        "header_location": header_location,
                        "via_target_location": None,
                        "action": display["action"],
                        "target": display["target"],
                        "command": display["block_command"],
                    }
                )

            headers.append(
                {
                    "file": filename,
                    "range_start": range_start,
                    "range_stop": range_stop,
                    "header_location": header_location,
                    "next_header": next_header,
                    "candidate_targets": subhandlers,
                    "inline_displays": inline_displays,
                    "dominant_action": action_counter.most_common(1)[0][0] if action_counter else "",
                    "dominant_target": target_counter.most_common(1)[0][0] if target_counter else "",
                    "text_contexts": text_contexts,
                }
            )

    return headers


def build_text_context_lookup(headers: list[dict[str, Any]]) -> dict[int, list[dict[str, Any]]]:
    lookup: dict[int, dict[tuple[Any, ...], dict[str, Any]]] = defaultdict(dict)

    for header in headers:
        for context in header["text_contexts"]:
            key = (
                context["header_location"],
                context["via_target_location"],
                context["action"],
                context["target"],
                context["command"],
            )
            lookup[context["text_offset"]][key] = context

    return {
        text_offset: sorted(
            context_map.values(),
            key=lambda item: (
                item["header_location"],
                -1 if item["via_target_location"] is None else item["via_target_location"],
                item["action"],
                item["target"],
                item["command"],
            ),
        )
        for text_offset, context_map in lookup.items()
    }


def scan_dispatch_context_lookup(
    filename: str,
    pos_exe: bytes,
    sheet_rows: list[dict[str, Any]],
) -> dict[int, list[dict[str, Any]]]:
    return build_text_context_lookup(scan_dispatch_headers(filename, pos_exe, sheet_rows))


def human_report(results: dict[str, list[dict[str, Any]]]) -> str:
    lines: list[str] = []
    for filename, headers in results.items():
        lines.append(filename)
        lines.append(f"  Headers: {len(headers)}")
        for header in headers:
            dominant_bits = []
            if header["dominant_action"]:
                dominant_bits.append(f"action={header['dominant_action']}")
            if header["dominant_target"]:
                dominant_bits.append(f"target={header['dominant_target']}")
            dominant_summary = ", ".join(dominant_bits) if dominant_bits else "no inferred context"

            lines.append(
                f"  {hex5(header['header_location'])} "
                f"(range {hex5(header['range_start'])}-{hex5(header['range_stop'])}, {dominant_summary})"
            )

            if header["inline_displays"]:
                lines.append("    Inline displays:")
                for display in header["inline_displays"]:
                    detail = f"arg={display['arg']}" if display["kind"] == "text" else f"rows={display['row_count']}"
                    label = display["block_command"] or "(no block command)"
                    lines.append(
                        f"      {hex5(display['opcode_location'])}: {display['kind']} "
                        f"{hex5(display['text_offset'])} {detail} -> {label}"
                    )

            if not header["candidate_targets"]:
                lines.append("    Candidate targets: none")
                continue

            lines.append("    Candidate targets:")
            for target in header["candidate_targets"]:
                inferred = []
                if target["inferred_action"]:
                    inferred.append(target["inferred_action"])
                if target["inferred_target"]:
                    inferred.append(target["inferred_target"])
                inferred_summary = " / ".join(inferred) if inferred else "unknown"
                lines.append(
                    f"      {hex5(target['table_location'])}: rel={hex5(target['relative_offset'])} "
                    f"-> {hex5(target['target_location'])} [{inferred_summary}]"
                )
                if target["flags"]:
                    flag_summary = ", ".join(
                        f"{flag['kind']}({hex(flag['arg1'])},{hex(flag['arg2'])}) @ {hex5(flag['location'])}"
                        for flag in target["flags"]
                    )
                    lines.append(f"        Flags: {flag_summary}")
                if target["displays"]:
                    for display in target["displays"]:
                        detail = f"arg={display['arg']}" if display["kind"] == "text" else f"rows={display['row_count']}"
                        label = display["block_command"] or "(no block command)"
                        lines.append(
                            f"        {hex5(display['opcode_location'])}: {display['kind']} "
                            f"{hex5(display['text_offset'])} {detail} -> {label}"
                        )
                else:
                    lines.append("        Displays: none found nearby")
        lines.append("")
    return "\n".join(lines).rstrip() + ("\n" if lines else "")


def json_ready(results: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for filename, headers in results.items():
        normalized_headers = []
        for header in headers:
            normalized_headers.append(
                {
                    "header_location": hex5(header["header_location"]),
                    "range_start": hex5(header["range_start"]),
                    "range_stop": hex5(header["range_stop"]),
                    "next_header": hex5(header["next_header"]) if header["next_header"] is not None else None,
                    "dominant_action": header["dominant_action"],
                    "dominant_target": header["dominant_target"],
                    "inline_displays": [
                        {
                            **display,
                            "opcode_location": hex5(display["opcode_location"]),
                            "text_offset": hex5(display["text_offset"]),
                        }
                        for display in header["inline_displays"]
                    ],
                    "candidate_targets": [
                        {
                            "table_location": hex5(target["table_location"]),
                            "relative_offset": hex5(target["relative_offset"]),
                            "target_location": hex5(target["target_location"]),
                            "inferred_action": target["inferred_action"],
                            "inferred_target": target["inferred_target"],
                            "flags": [
                                {
                                    **flag,
                                    "location": hex5(flag["location"]),
                                    "arg1": hex(flag["arg1"]),
                                    "arg2": hex(flag["arg2"]),
                                    "extra": hex(flag["extra"]) if "extra" in flag else None,
                                }
                                for flag in target["flags"]
                            ],
                            "displays": [
                                {
                                    **display,
                                    "opcode_location": hex5(display["opcode_location"]),
                                    "text_offset": hex5(display["text_offset"]),
                                }
                                for display in target["displays"]
                            ],
                        }
                        for target in header["candidate_targets"]
                    ],
                    "text_contexts": [
                        {
                            **context,
                            "text_offset": hex5(context["text_offset"]),
                            "opcode_location": hex5(context["opcode_location"]),
                            "header_location": hex5(context["header_location"]),
                            "via_target_location": hex5(context["via_target_location"])
                            if context["via_target_location"] is not None
                            else None,
                        }
                        for context in header["text_contexts"]
                    ],
                }
            )
        output[filename] = normalized_headers
    return output


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Analyze POS.EXE room/action dispatch headers and map reachable text back to workbook commands."
    )
    parser.add_argument("--files", nargs="+", default=["POS1.MSD", "YUMI.MSD"])
    parser.add_argument("--json", action="store_true", help="Output JSON instead of human-readable text")
    args = parser.parse_args()

    file_filter = set(args.files)
    dump_rows = load_dump_rows(Path(DUMP_XLS_PATH), file_filter)
    pos_exe = (Path(ORIGINAL_ROM_DIR) / "POS.EXE").read_bytes()

    results = {
        filename: scan_dispatch_headers(filename, pos_exe, dump_rows.get(filename, []))
        for filename in args.files
    }

    if args.json:
        print(json.dumps(json_ready(results), indent=2))
    else:
        print(human_report(results))


if __name__ == "__main__":
    main()
