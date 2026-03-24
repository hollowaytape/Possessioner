from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rominfo import ARRIVAL_POINTERS, DUMP_XLS_PATH, ORIGINAL_ROM_DIR, POINTER_DUMP_XLS_PATH


TOKEN_RE = re.compile(r"\[([^\]]+)\]")


def parse_hex(value: str | None) -> int | None:
    if value is None:
        return None
    value = str(value).strip()
    if not value:
        return None
    return int(value, 16)


def hex5(value: int | None) -> str | None:
    if value is None:
        return None
    return f"0x{value:05x}"


def bytes_to_hex(data: bytes) -> str:
    return " ".join(f"{b:02x}" for b in data)


def tokenize_ctrl_codes(ctrl_codes: str | None) -> list[str]:
    if not ctrl_codes:
        return []
    return TOKEN_RE.findall(ctrl_codes)


def parse_dialogue_metadata(ctrl_codes: str | None) -> dict[str, Any]:
    tokens = tokenize_ctrl_codes(ctrl_codes)
    metadata: dict[str, Any] = {
        "tokens": tokens,
        "speaker": None,
        "portrait": None,
        "text_color": None,
        "flow": None,
    }

    color_tokens = {"Black", "Blue", "Red", "Purple", "Green", "Cyan", "Yellow", "White"}
    portrait_states = {"Neutral", "Energetic", "Upset", "Surprised", "Happy", "Sad", "Excited"}
    flow_tokens = {"Start", "Continue", "Narration"}
    default_speaker_colors = {
        "Alisa": "White",
        "Honghua": "Cyan",
        "Meryl": "Green",
        "Nedra": "Purple",
        "Kumiko": "Yellow",
        "Yumi": "Yellow",
        "Ayaka": "Yellow",
        "Misha": "Yellow",
        "Prim": "Yellow",
        "Doctor": "Yellow",
        "Doc": "Yellow",
        "Eris": "Green",
        "Deal": "Cyan",
        "Mechanic1": "Cyan",
        "Mechanic2": "Cyan",
        "Mechanic3": "Cyan",
        "Carmine": "Cyan",
        "Rashmar": "Cyan",
        "Clerk": "White",
        "Owner": "Cyan",
        "Master": "Cyan",
        "Assistant": "Cyan",
        "Person": "Cyan",
        "Tina": "Yellow",
        "May": "Yellow",
        "Iris": "Yellow",
        "Fairy": "Yellow",
        "Possessioner": "White",
    }

    for token in tokens:
        if token in color_tokens and metadata["text_color"] is None:
            metadata["text_color"] = token
            continue
        if token in flow_tokens and metadata["flow"] is None:
            metadata["flow"] = token
            continue
        if token == "Narration":
            metadata["speaker"] = "Narration"
            metadata["flow"] = "Narration"
            continue
        if "-" in token:
            left, right = token.rsplit("-", 1)
            if right in color_tokens:
                metadata["speaker"] = left
                metadata["text_color"] = right
                continue
            if right in portrait_states:
                metadata["speaker"] = left
                metadata["portrait"] = right
                continue
            if right in {"Start", "Continue"}:
                metadata["speaker"] = left
                metadata["flow"] = right
                continue
        if metadata["speaker"] is None:
            metadata["speaker"] = token

    if metadata["speaker"] and metadata["text_color"] is None:
        metadata["text_color"] = default_speaker_colors.get(metadata["speaker"])

    return metadata


@dataclass
class ParsedCommand:
    kind: str
    start: int
    length: int
    raw: str
    confidence: str
    details: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            "kind": self.kind,
            "start": hex5(self.start),
            "length": self.length,
            "raw": self.raw,
            "confidence": self.confidence,
            "details": self.details,
        }


def parse_known_command(data: bytes, start: int, current_ptr_loc: int | None = None) -> ParsedCommand | None:
    if start < 0 or start >= len(data):
        return None

    opcode = data[start]

    if opcode == 0x02 and start + 3 < len(data) and data[start + 1] != 0xFF:
        lo = data[start + 1]
        hi = data[start + 2]
        arg = data[start + 3]
        text_location = (hi << 8) | lo
        details: dict[str, Any] = {
            "opcode": "0x02",
            "text_location": hex5(text_location),
            "arg": arg,
        }
        if current_ptr_loc is not None:
            details["matches_pointer_location"] = current_ptr_loc == start + 1
        return ParsedCommand(
            kind="text",
            start=start,
            length=4,
            raw=bytes_to_hex(data[start : start + 4]),
            confidence="high",
            details=details,
        )

    if opcode == 0x02 and start + 4 < len(data) and data[start + 1] == 0xFF:
        subtype = data[start + 2]
        arg1 = data[start + 3]
        arg2 = data[start + 4]
        meaning = "extended_02ff"
        confidence = "medium"
        if subtype == 0x01:
            meaning = "set_flag"
        elif subtype == 0x02:
            meaning = "check_flag"
        elif arg2 == 0x00:
            meaning = "possible_flag_or_branch"
            confidence = "low"
        return ParsedCommand(
            kind=meaning,
            start=start,
            length=5,
            raw=bytes_to_hex(data[start : start + 5]),
            confidence=confidence,
            details={
                "opcode": "0x02ff",
                "subtype": f"0x{subtype:02x}",
                "arg1": f"0x{arg1:02x}",
                "arg2": f"0x{arg2:02x}",
            },
        )

    if opcode == 0x03 and start + 5 < len(data) and data[start + 1] == 0xFF:
        payload = data[start + 2 : start + 6]
        meaning = "extended_03ff"
        if payload[0] == 0x01:
            meaning = "clear_flag"
        return ParsedCommand(
            kind=meaning,
            start=start,
            length=6,
            raw=bytes_to_hex(data[start : start + 6]),
            confidence="medium" if meaning == "clear_flag" else "low",
            details={
                "opcode": "0x03ff",
                "payload": [f"0x{b:02x}" for b in payload],
            },
        )

    if opcode == 0x01 and start + 3 < len(data) and data[start + 3] == 0xFF:
        value = (data[start + 2] << 8) | data[start + 1]
        return ParsedCommand(
            kind="possible_room_or_action_selector",
            start=start,
            length=4,
            raw=bytes_to_hex(data[start : start + 4]),
            confidence="low",
            details={
                "opcode": "0x01",
                "value": hex5(value),
            },
        )

    return None


def parse_context_window(data: bytes, pointer_location: int, window: int = 24) -> dict[str, Any]:
    command_start = max(pointer_location - 1, 0)
    window_start = max(command_start - window, 0)
    ops: list[dict[str, Any]] = []
    cursor = window_start

    while cursor < command_start:
        parsed = parse_known_command(data, cursor)
        if parsed is not None and cursor + parsed.length <= command_start:
            ops.append(parsed.to_dict())
            cursor += parsed.length
            continue

        unknown_start = cursor
        cursor += 1
        while cursor < command_start:
            if parse_known_command(data, cursor) is not None:
                break
            cursor += 1
        raw = data[unknown_start:cursor]
        ops.append(
            {
                "kind": "unknown",
                "start": hex5(unknown_start),
                "length": len(raw),
                "raw": bytes_to_hex(raw),
                "confidence": "unknown",
                "details": {},
            }
        )

    current = parse_known_command(data, command_start, current_ptr_loc=pointer_location)
    current_dict = current.to_dict() if current is not None else None
    next_bytes = data[command_start : min(command_start + 12, len(data))]

    return {
        "window_start": hex5(window_start),
        "window_end": hex5(min(command_start + 12, len(data))),
        "window_bytes": bytes_to_hex(data[window_start : min(command_start + 12, len(data))]),
        "preceding_ops": ops,
        "current_command": current_dict,
        "next_bytes": bytes_to_hex(next_bytes),
    }


def load_dump_rows(path: Path) -> dict[str, dict[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    result: dict[str, dict[int, dict[str, Any]]] = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(rows)]
        except StopIteration:
            continue

        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        if "Offset" not in name_to_index or "Japanese" not in name_to_index:
            continue

        sheet_rows: dict[int, dict[str, Any]] = {}
        for row in rows:
            offset = parse_hex(row[name_to_index["Offset"]])
            japanese = row[name_to_index["Japanese"]]
            if offset is None or japanese is None:
                continue

            sheet_rows[offset] = {
                "offset": hex5(offset),
                "command": row[name_to_index.get("Command", -1)] if "Command" in name_to_index else None,
                "ctrl_codes": row[name_to_index.get("Ctrl Codes", -1)] if "Ctrl Codes" in name_to_index else None,
                "japanese": japanese,
                "english": row[name_to_index.get("English", -1)] if "English" in name_to_index else None,
                "english_typeset": row[name_to_index.get("English (Typeset)", -1)] if "English (Typeset)" in name_to_index else None,
                "comments": row[name_to_index.get("Comments", -1)] if "Comments" in name_to_index else None,
            }
        result[sheet_name] = sheet_rows

    return result


def load_pointer_rows(path: Path) -> dict[str, dict[int, list[dict[str, Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    result: dict[str, dict[int, list[dict[str, Any]]]] = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(rows)]
        except StopIteration:
            continue

        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        if "Text Loc" not in name_to_index or "Ptr Loc" not in name_to_index:
            continue

        sheet_rows: dict[int, list[dict[str, Any]]] = {}
        for row in rows:
            text_loc = parse_hex(row[name_to_index["Text Loc"]])
            ptr_loc = parse_hex(row[name_to_index["Ptr Loc"]])
            if text_loc is None or ptr_loc is None:
                continue

            sheet_rows.setdefault(text_loc, []).append(
                {
                    "pointer_location": hex5(ptr_loc),
                    "pointer_location_int": ptr_loc,
                    "bytes_preview": row[name_to_index.get("Bytes", -1)] if "Bytes" in name_to_index else None,
                    "points_to": row[name_to_index.get("Points To", -1)] if "Points To" in name_to_index else None,
                    "comments": row[name_to_index.get("Comments", -1)] if "Comments" in name_to_index else None,
                }
            )
        result[sheet_name] = sheet_rows

    return result


def infer_pointer_heuristics(
    record: dict[str, Any],
    pointer_entry: dict[str, Any],
    context: dict[str, Any],
) -> dict[str, Any]:
    labels: list[str] = []
    current = context.get("current_command")
    preceding_ops = context.get("preceding_ops", [])
    recognized_ops = [op for op in preceding_ops if op["kind"] != "unknown"]
    preceding_text_ops = [op for op in recognized_ops if op["kind"] == "text"]

    command_name = record.get("command")
    if command_name:
        labels.append("named-workbook-command")
        if command_name.startswith("Talk"):
            labels.append("talk-action")
        elif command_name.startswith("Look"):
            labels.append("look-action")
        elif command_name == "Think":
            labels.append("think-action")
        elif command_name == "(Arrive)":
            labels.append("arrival-action")
        elif command_name == "(continued)":
            labels.append("continued-window")
        elif command_name == "?":
            labels.append("unknown-workbook-command")

    if record.get("offset") == "0x00000":
        labels.append("offset-zero-row")
    if pointer_entry.get("comments") == "Arrival pointer override from rominfo.py":
        labels.append("arrival-pointer-override")

    if current is not None and current["kind"] == "text":
        labels.append("direct-text-command")
        arg = current["details"].get("arg")
        labels.append(f"text-arg-{arg}")

    if any(op["kind"] == "check_flag" for op in recognized_ops):
        labels.append("flag-gated")
    if any(op["kind"] == "set_flag" for op in recognized_ops):
        labels.append("flag-mutating")
    if any(op["kind"] == "clear_flag" for op in recognized_ops):
        labels.append("flag-clearing")
    if any(op["kind"] in ("possible_flag_or_branch", "extended_03ff", "possible_room_or_action_selector") for op in recognized_ops):
        labels.append("branchy-context")
    if preceding_text_ops:
        labels.append("text-chain")

    metadata = record.get("dialogue_metadata", {})
    if metadata.get("speaker"):
        labels.append("speaker-known")
    if metadata.get("portrait"):
        labels.append("portrait-known")
    if metadata.get("flow") == "Narration":
        labels.append("narration")

    previous_text_locations = [op["details"].get("text_location") for op in preceding_text_ops if op["details"].get("text_location")]
    return {
        "labels": sorted(set(labels)),
        "previous_text_locations": previous_text_locations,
        "recognized_op_kinds": [op["kind"] for op in recognized_ops],
        "current_text_arg": current["details"].get("arg") if current and current["kind"] == "text" else None,
    }


def infer_record_labels(record: dict[str, Any]) -> list[str]:
    labels: set[str] = set()
    metadata = record.get("dialogue_metadata", {})
    if record.get("pointer_count", 0) == 0:
        labels.add("no-pointer-context")
    if record.get("pointer_count", 0) > 1:
        labels.add("multi-pointer-row")
    if metadata.get("speaker"):
        labels.add("speaker-known")
    if record.get("english"):
        labels.add("translated")
    if record.get("english_typeset"):
        labels.add("typeset")

    for pointer_context in record.get("pointer_contexts", []):
        for label in pointer_context.get("heuristics", {}).get("labels", []):
            labels.add(label)

    return sorted(labels)


def build_records(
    dump_rows: dict[str, dict[int, dict[str, Any]]],
    pointer_rows: dict[str, dict[int, list[dict[str, Any]]]],
    pos_exe_bytes: bytes,
    file_filter: str | None = None,
    offset_filter: int | None = None,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    for sheet_name, rows in dump_rows.items():
        if not sheet_name.endswith(".MSD"):
            continue
        if file_filter and sheet_name != file_filter:
            continue

        file_pointers = pointer_rows.get(sheet_name, {})
        arrival_pointers = ARRIVAL_POINTERS.get(sheet_name, [])

        for offset in sorted(rows):
            if offset_filter is not None and offset != offset_filter:
                continue
            row = rows[offset]
            pointer_entries = list(file_pointers.get(offset, []))

            if offset == 0:
                for _, pointer_location in arrival_pointers:
                    pointer_entries.append(
                        {
                            "pointer_location": hex5(pointer_location),
                            "pointer_location_int": pointer_location,
                            "bytes_preview": "Arrival text",
                            "points_to": None,
                            "comments": "Arrival pointer override from rominfo.py",
                        }
                    )

            pointer_contexts = []
            for entry in pointer_entries:
                pointer_loc = entry["pointer_location_int"]
                context = parse_context_window(pos_exe_bytes, pointer_loc)
                heuristics = infer_pointer_heuristics(row, entry, context)
                pointer_contexts.append(
                    {
                        "pointer_location": entry["pointer_location"],
                        "bytes_preview": entry["bytes_preview"],
                        "points_to": entry["points_to"],
                        "comments": entry["comments"],
                        "context": context,
                        "heuristics": heuristics,
                    }
                )

            record = {
                "file": sheet_name,
                "offset": row["offset"],
                "command": row["command"],
                "ctrl_codes": row["ctrl_codes"],
                "dialogue_metadata": parse_dialogue_metadata(row["ctrl_codes"]),
                "japanese": row["japanese"],
                "english": row["english"],
                "english_typeset": row["english_typeset"],
                "comments": row["comments"],
                "pointer_count": len(pointer_contexts),
                "pointer_contexts": pointer_contexts,
            }
            record["labels"] = infer_record_labels(record)
            records.append(record)

    return records


def summarize(records: list[dict[str, Any]]) -> dict[str, Any]:
    by_file: dict[str, dict[str, int]] = {}
    parsed_pointer_count = 0
    total_pointer_count = 0

    for record in records:
        file_summary = by_file.setdefault(record["file"], {"rows": 0, "rows_with_pointers": 0, "pointers": 0})
        file_summary["rows"] += 1
        if record["pointer_count"] > 0:
            file_summary["rows_with_pointers"] += 1
            file_summary["pointers"] += record["pointer_count"]

        total_pointer_count += record["pointer_count"]
        for pointer_context in record["pointer_contexts"]:
            if pointer_context["context"]["current_command"] is not None:
                parsed_pointer_count += 1

    return {
        "record_count": len(records),
        "total_pointer_count": total_pointer_count,
        "parsed_pointer_count": parsed_pointer_count,
        "files": by_file,
    }


def describe_op(op: dict[str, Any]) -> str:
    kind = op["kind"]
    start = op["start"]
    details = op.get("details", {})

    if kind == "text":
        text_location = details.get("text_location")
        arg = details.get("arg")
        suffix = " <- current row" if details.get("matches_pointer_location") else ""
        return f"{start}: show text at {text_location} (arg={arg}){suffix}"

    if kind == "check_flag":
        return (
            f"{start}: check flag "
            f"(flag_offset={details.get('arg1')}, bit_index={details.get('arg2')})"
        )

    if kind == "set_flag":
        return (
            f"{start}: set flag "
            f"(flag_offset={details.get('arg1')}, bit_index={details.get('arg2')})"
        )

    if kind == "clear_flag":
        payload = details.get("payload", [])
        if len(payload) >= 3:
            return f"{start}: clear flag (flag_offset={payload[1]}, bit_index={payload[2]})"
        return f"{start}: clear flag-like op (payload={', '.join(payload)})"

    if kind == "extended_03ff":
        payload = ", ".join(details.get("payload", []))
        return f"{start}: unknown 03 ff op; likely non-text script command (payload={payload})"

    if kind == "possible_room_or_action_selector":
        return f"{start}: possible selector/value op (value={details.get('value')})"

    if kind == "unknown":
        return f"{start}: unknown bytes [{op['raw']}]"

    return f"{start}: {kind} [{op['raw']}]"


def render_human(records: list[dict[str, Any]]) -> str:
    lines: list[str] = []

    for record in records:
        metadata = record["dialogue_metadata"]
        speaker_bits = []
        if metadata.get("speaker"):
            speaker_bits.append(metadata["speaker"])
        if metadata.get("portrait"):
            speaker_bits.append(metadata["portrait"])
        if metadata.get("text_color"):
            speaker_bits.append(metadata["text_color"])
        if metadata.get("flow"):
            speaker_bits.append(metadata["flow"])
        speaker_summary = ", ".join(speaker_bits) if speaker_bits else "unknown"

        lines.append(f"{record['file']} {record['offset']}")
        lines.append(f"  Workbook command: {record['command'] or '(none)'}")
        lines.append(f"  Dialogue metadata: {speaker_summary}")
        if record.get("labels"):
            lines.append(f"  Heuristic labels: {', '.join(record['labels'])}")
        if record["ctrl_codes"]:
            lines.append(f"  Ctrl codes: {record['ctrl_codes']}")
        lines.append(f"  Japanese: {record['japanese']}")
        if record["english"]:
            lines.append(f"  English: {record['english']}")
        if record["english_typeset"]:
            lines.append(f"  Typeset: {record['english_typeset']}")
        if record["comments"]:
            lines.append(f"  Comments: {record['comments']}")

        if not record["pointer_contexts"]:
            lines.append("  Pointer context: none found")
            lines.append("")
            continue

        lines.append(f"  Pointer contexts: {record['pointer_count']}")
        for idx, pointer_context in enumerate(record["pointer_contexts"], start=1):
            context = pointer_context["context"]
            lines.append(f"    [{idx}] Ptr loc {pointer_context['pointer_location']}")
            if pointer_context.get("comments"):
                lines.append(f"        Notes: {pointer_context['comments']}")
            heuristic_labels = pointer_context.get("heuristics", {}).get("labels", [])
            if heuristic_labels:
                lines.append(f"        Heuristic labels: {', '.join(heuristic_labels)}")
            previous_text_locations = pointer_context.get("heuristics", {}).get("previous_text_locations", [])
            if previous_text_locations:
                lines.append(f"        Previous text targets nearby: {', '.join(previous_text_locations)}")
            current = context.get("current_command")
            if current is not None:
                lines.append(f"        Current command: {describe_op(current)}")
            else:
                lines.append("        Current command: unparsed")

            preceding_ops = [op for op in context.get("preceding_ops", []) if op["kind"] != "unknown"]
            if preceding_ops:
                lines.append("        Nearby recognized ops:")
                for op in preceding_ops[-5:]:
                    lines.append(f"          - {describe_op(op)}")
            else:
                lines.append("        Nearby recognized ops: none")

            unknown_ops = [op for op in context.get("preceding_ops", []) if op["kind"] == "unknown"]
            if unknown_ops:
                lines.append(f"        Unknown byte runs nearby: {len(unknown_ops)}")
            lines.append(f"        Raw window: {context['window_bytes']}")
        lines.append("")

    return "\n".join(lines).rstrip() + ("\n" if lines else "")


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze MSD dialog rows with nearby POS.EXE pointer-command context.")
    parser.add_argument("--file", help="Analyze a single MSD sheet, e.g. P_HON1.MSD")
    parser.add_argument("--offset", help="Analyze a single text offset, e.g. 0x00163")
    parser.add_argument("--limit", type=int, help="Limit the number of output records")
    parser.add_argument("--output", help="Write detailed JSON to a file instead of stdout")
    parser.add_argument("--summary-only", action="store_true", help="Print only the summary")
    parser.add_argument("--human", action="store_true", help="Print a human-readable explanation instead of JSON")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON output")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    dump_rows = load_dump_rows(root / DUMP_XLS_PATH)
    pointer_rows = load_pointer_rows(root / POINTER_DUMP_XLS_PATH)
    pos_exe_bytes = (root / ORIGINAL_ROM_DIR / "POS.EXE").read_bytes()
    offset_filter = parse_hex(args.offset)

    records = build_records(
        dump_rows,
        pointer_rows,
        pos_exe_bytes,
        file_filter=args.file,
        offset_filter=offset_filter,
    )
    if args.limit is not None:
        records = records[: args.limit]

    summary = summarize(records)
    if args.human:
        text = render_human(records)
        if args.output:
            Path(args.output).write_text(text, encoding="utf-8")
            print(f"Wrote {len(records)} human-readable records to {args.output}")
        else:
            print(text)
        return

    payload: dict[str, Any] = {"summary": summary}
    if not args.summary_only:
        payload["records"] = records

    text = json.dumps(payload, ensure_ascii=True, indent=2 if args.pretty else None)
    if args.output:
        Path(args.output).write_text(text, encoding="utf-8")
        print(f"Wrote {len(records)} records to {args.output}")
    else:
        print(text)


if __name__ == "__main__":
    main()
