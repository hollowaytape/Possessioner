from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rominfo import DUMP_XLS_PATH

SECTION_PATTERN = re.compile(r"<h2[^>]*>([^<]+).*?</h2>\s*<div>\s*(.*?)\s*</div>", re.DOTALL)
TAG_PATTERN = re.compile(r"<[^>]+>")
PAREN_PATTERN = re.compile(r"\([^)]*\)|\uff08[^\uff09]*\uff09")
NON_WORD_PATTERN = re.compile(r"[^a-z0-9]+")
REPEAT_PATTERN = re.compile(r"\s*[x×]\s*(\d+)\s*$", re.IGNORECASE)

ACTION_MAP = {
    "見る": "Look",
    "話す": "Talk",
    "移動": "Move",
    "調べる": "Examine",
    "思う": "Think",
    "考える": "Think",
    "使う": "Use",
    "浴びる": "Shower",
    "脱ぐ": "Undress",
    "昇る": "Climb",
    "開ける": "Open",
    "聞く": "Listen",
    "乱入する": "Interrupt",
}

TARGET_MAP = {
    "ホンホア": "Honghua",
    "メリル": "Meryl",
    "ネドラ": "Nedra",
    "オペレーター": "Operators",
    "店員": "Clerk",
    "オーナー": "Owner",
    "主人": "Master",
    "みんな": "Everyone",
    "まわり": "Around",
    "通行人": "Passerby",
    "ユミ＝中山": "Yumi",
    "ポゼッショナー": "Possessioner",
    "出現現場": "Appearance Site",
    "廊下": "Corridor",
    "控室": "Lounge",
    "シャワールーム": "Shower Room",
    "シャワー": "Shower",
    "街": "City",
    "商業地区": "Business Zone",
    "多層構造地区": "Central Hub",
    "旧市街": "Old City",
    "旧市街地": "Old City",
    "放棄地区": "Abandoned Zone",
    "廃ビル": "Abandoned Building",
    "エレベーター": "Elevator",
    "部屋": "Room",
    "ミシャス": "Michass",
    "整備室": "Maint Room",
    "階段": "Stairs",
    "3人": "Three",
    "二人": "Two",
    "ジャンク屋": "Junk Shop",
    "病棟": "Medical Ward",
    "ドクター": "Doctor",
    "ラボテック": "Labotech",
    "上": "Up",
    "道路": "Road",
    "発信地点": "Signal Point",
    "中に入る": "Enter",
    "前に進む": "Forward",
    "女帝": "Empress",
    "ボックス": "Box",
    "下水入口": "Sewer Entrance",
    "下水の中へ": "Into Sewer",
    "メイ": "May",
    "男": "Man",
    "ラシュマル": "Rashmar",
    "コンピューター": "Computer",
    "床": "Floor",
    "ドア": "Door",
    "資料室": "Archives",
    "整備員": "Crew",
    "制御室": "Control Room",
    "制御卓": "Console",
    "奥に進む": "Forward",
    "右": "Right",
    "左": "Left",
    "真っすぐ": "Straight",
    "前": "Forward",
    "ティナ": "Tina",
    "妖精": "Fairy",
    "外に出る": "Exit",
}

COMMAND_ACTIONS = sorted(
    {
        "Look",
        "Talk",
        "Move",
        "Examine",
        "Think",
        "Use",
        "Climb",
        "Open",
        "Listen",
        "Shower",
        "Interrupt",
        "Undress",
        "Board",
        "Enter",
    },
    key=len,
    reverse=True,
)

STOPWORDS = {
    "",
    "the",
    "a",
    "an",
    "of",
    "to",
    "and",
    "after",
    "before",
    "while",
    "have",
    "has",
    "had",
    "just",
    "still",
    "here",
    "there",
    "later",
    "attack",
    "battle",
    "scene",
    "visit",
    "talking",
    "looking",
    "heading",
    "points",
    "pointed",
    "out",
    "yells",
    "about",
    "from",
    "again",
    "everyone",
    "all",
}


@dataclass
class Step:
    raw: str
    action_jp: str | None
    target_jp: str | None
    action_en: str | None
    target_en: str | None
    repeat: int
    notes: list[str]


def clean_text(value: str) -> str:
    text = value.replace("&nbsp;", " ").replace("\u3000", " ")
    text = TAG_PATTERN.sub("", text)
    return re.sub(r"\s+", " ", text).strip()


def split_notes(text: str) -> tuple[str, list[str]]:
    notes: list[str] = []
    for pattern in (r"\(([^)]*)\)", r"\uff08([^\uff09]*)\uff09"):
        while True:
            match = re.search(pattern, text)
            if not match:
                break
            notes.append(match.group(1).strip())
            text = text[: match.start()] + text[match.end() :]
    return re.sub(r"\s+", " ", text).strip(), notes


def parse_step(line: str) -> Step:
    stripped = clean_text(line)
    stripped, notes = split_notes(stripped)
    repeat = 1
    repeat_match = REPEAT_PATTERN.search(stripped)
    if repeat_match:
        repeat = int(repeat_match.group(1))
        stripped = stripped[: repeat_match.start()].strip()

    action_jp = None
    target_jp = None
    if " " in stripped:
        action_jp, target_jp = stripped.split(" ", 1)
    else:
        action_jp = stripped

    action_en = ACTION_MAP.get(action_jp or "")
    target_en = translate_target(target_jp) if target_jp else None
    return Step(
        raw=stripped,
        action_jp=action_jp,
        target_jp=target_jp,
        action_en=action_en,
        target_en=target_en,
        repeat=repeat,
        notes=notes,
    )


def translate_target(text: str | None) -> str | None:
    if not text:
        return None
    parts = [part.strip() for part in text.split("→")]
    translated_parts = []
    for part in parts:
        translated_parts.append(TARGET_MAP.get(part, part))
    return " -> ".join(translated_parts)


def parse_sections(path: Path) -> list[dict[str, Any]]:
    html = path.read_text(encoding="utf-8", errors="ignore")
    sections: list[dict[str, Any]] = []
    for match in SECTION_PATTERN.finditer(html):
        heading = clean_text(match.group(1))
        body = match.group(2).replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
        lines = [clean_text(line) for line in body.splitlines()]
        steps = [parse_step(line) for line in lines if clean_text(line)]
        if not steps:
            continue
        sections.append(
            {
                "heading_jp": heading,
                "heading_en": TARGET_MAP.get(heading, heading),
                "steps": [
                    {
                        "raw": step.raw,
                        "action_jp": step.action_jp,
                        "target_jp": step.target_jp,
                        "action_en": step.action_en,
                        "target_en": step.target_en,
                        "repeat": step.repeat,
                        "notes": step.notes,
                    }
                    for step in steps
                ],
            }
        )
    return sections


def normalize_text(text: str | None) -> str:
    if not text:
        return ""
    text = PAREN_PATTERN.sub(" ", text)
    text = text.replace("maint.", "maint")
    text = text.replace("armd.", "armored")
    text = text.replace("michaas", "michass")
    text = text.replace("all", "everyone")
    text = NON_WORD_PATTERN.sub(" ", text.lower())
    tokens = [token for token in text.split() if token not in STOPWORDS and not token.isdigit()]
    return " ".join(tokens)


def parse_command(command: str) -> tuple[str | None, str]:
    text = str(command).strip()
    for action in COMMAND_ACTIONS:
        if text == action:
            return action, ""
        prefix = f"{action} "
        prefix_dash = f"{action} - "
        if text.startswith(prefix_dash):
            return action, normalize_text(text[len(prefix_dash) :])
        if text.startswith(prefix):
            return action, normalize_text(text[len(prefix) :])
    return None, normalize_text(text)


def load_sheet_commands(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    commands_by_sheet: dict[str, list[dict[str, Any]]] = {}
    try:
        for sheet_name in workbook.sheetnames:
            if not sheet_name.endswith(".MSD"):
                continue
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = [cell for cell in next(rows)]
            except StopIteration:
                continue
            name_to_index = {name: index for index, name in enumerate(header) if name is not None}
            if "Command" not in name_to_index or "Offset" not in name_to_index:
                continue
            sheet_commands: list[dict[str, Any]] = []
            for row in rows:
                command = row[name_to_index["Command"]]
                if not command:
                    continue
                action, normalized_target = parse_command(str(command))
                sheet_commands.append(
                    {
                        "offset": row[name_to_index["Offset"]],
                        "command": str(command).strip(),
                        "action": action,
                        "target": normalized_target,
                    }
                )
            commands_by_sheet[sheet_name] = sheet_commands
    finally:
        workbook.close()
    return commands_by_sheet


def step_matches_command(step: dict[str, Any], command: dict[str, Any]) -> bool:
    action = step.get("action_en")
    if not action or action != command.get("action"):
        return False
    if action == "Move":
        target = normalize_text(step.get("target_en"))
        return bool(target) and target in command.get("target", "")

    step_target = normalize_text(step.get("target_en"))
    command_target = command.get("target", "")
    if not step_target:
        return command_target == ""
    if step_target == command_target:
        return True
    step_tokens = set(step_target.split())
    command_tokens = set(command_target.split())
    if not step_tokens or not command_tokens:
        return False
    overlap = step_tokens & command_tokens
    return len(overlap) >= min(2, len(step_tokens), len(command_tokens))


def score_section(section: dict[str, Any], commands_by_sheet: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    steps = section["steps"]
    for sheet_name, commands in commands_by_sheet.items():
        match_count = 0
        matched_steps: list[dict[str, Any]] = []
        for index, step in enumerate(steps):
            matches = [command for command in commands if step_matches_command(step, command)]
            if not matches:
                continue
            match_count += 1
            matched_steps.append(
                {
                    "step_index": index,
                    "step": step["raw"],
                    "repeat": step["repeat"],
                    "matches": [match["command"] for match in matches[:5]],
                }
            )
        if not match_count:
            continue
        results.append(
            {
                "sheet": sheet_name,
                "matched_step_count": match_count,
                "step_count": len(steps),
                "coverage": round(match_count / len(steps), 3),
                "matched_steps": matched_steps[:10],
            }
        )
    results.sort(key=lambda item: (-item["matched_step_count"], -item["coverage"], item["sheet"]))
    return results[:5]


def build_report(walkthrough_path: Path, dump_path: Path) -> dict[str, Any]:
    sections = parse_sections(walkthrough_path)
    commands_by_sheet = load_sheet_commands(dump_path)
    enriched_sections: list[dict[str, Any]] = []
    for index, section in enumerate(sections):
        enriched_sections.append(
            {
                "index": index,
                **section,
                "candidate_sheets": score_section(section, commands_by_sheet),
            }
        )
    return {
        "walkthrough": str(walkthrough_path),
        "section_count": len(enriched_sections),
        "sections": enriched_sections,
    }


def human_report(report: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append(f"Walkthrough sections: {report['section_count']}")
    lines.append("")
    for section in report["sections"]:
        lines.append(f"[{section['index']:02d}] {section['heading_jp']} / {section['heading_en']}")
        for step in section["steps"]:
            rendered = step["raw"]
            if step["repeat"] > 1:
                rendered += f" x{step['repeat']}"
            if step["action_en"] or step["target_en"]:
                rendered += f" :: {step['action_en'] or '?'}"
                if step["target_en"]:
                    rendered += f" / {step['target_en']}"
            if step["notes"]:
                rendered += f" [{'; '.join(step['notes'])}]"
            lines.append(f"  - {rendered}")
        if section["candidate_sheets"]:
            lines.append("  Candidates:")
            for candidate in section["candidate_sheets"]:
                lines.append(
                    f"    - {candidate['sheet']} matched {candidate['matched_step_count']}/{candidate['step_count']}"
                )
        else:
            lines.append("  Candidates: (none)")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse docs\\walkthrough.htm and score likely workbook sheets per section.")
    parser.add_argument(
        "--walkthrough",
        type=Path,
        default=Path("docs") / "walkthrough.htm",
        help="Path to the walkthrough HTML export",
    )
    parser.add_argument(
        "--dump",
        type=Path,
        default=Path(DUMP_XLS_PATH),
        help="Path to PSSR_dump.xlsx",
    )
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of a human-readable report")
    args = parser.parse_args()

    report = build_report(args.walkthrough, args.dump)
    if args.json:
        print(json.dumps(report, indent=2, ensure_ascii=False))
    else:
        print(human_report(report))


if __name__ == "__main__":
    main()
