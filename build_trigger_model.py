from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from analyze_handoff_contexts import parse_offset
from rominfo import DUMP_XLS_PATH

DISPLAY_TYPE_HEADER = "Command Display Type"
DISPLAY_DETAIL_HEADER = "Command Display Detail"
DISPATCH_DETAIL_HEADER = "Command Dispatch Detail"
HANDOFF_DETAIL_HEADER = "Command Handoff Detail"


def hex5(value: int | None) -> str | None:
    if value is None:
        return None
    return f"0x{value:05x}"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def load_workbook_rows(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows_by_file: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith(".MSD"):
            continue
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(iterator)]
        except StopIteration:
            continue

        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        if "Offset" not in name_to_index:
            continue

        current_command = ""
        rows: list[dict[str, Any]] = []
        for row in iterator:
            offset = parse_offset(row[name_to_index["Offset"]])
            if offset is None:
                continue

            command = row[name_to_index["Command"]] if "Command" in name_to_index else None
            if command:
                current_command = str(command).strip()

            rows.append(
                {
                    "offset": offset,
                    "offset_hex": hex5(offset),
                    "command": str(command).strip() if command else "",
                    "block_command": current_command,
                    "english": row[name_to_index["English"]] if "English" in name_to_index else None,
                    "display_type": (
                        row[name_to_index[DISPLAY_TYPE_HEADER]]
                        if DISPLAY_TYPE_HEADER in name_to_index
                        else None
                    ),
                    "display_detail": (
                        row[name_to_index[DISPLAY_DETAIL_HEADER]]
                        if DISPLAY_DETAIL_HEADER in name_to_index
                        else None
                    ),
                    "dispatch_detail": (
                        row[name_to_index[DISPATCH_DETAIL_HEADER]]
                        if DISPATCH_DETAIL_HEADER in name_to_index
                        else None
                    ),
                    "handoff_detail": (
                        row[name_to_index[HANDOFF_DETAIL_HEADER]]
                        if HANDOFF_DETAIL_HEADER in name_to_index
                        else None
                    ),
                }
            )
        rows_by_file[sheet_name] = rows

    workbook.close()
    return rows_by_file


def build_dispatch_lookup(dispatch_data: dict[str, list[dict[str, Any]]]) -> dict[tuple[str, int], list[dict[str, Any]]]:
    lookup: dict[tuple[str, int], list[dict[str, Any]]] = {}
    for filename, headers in dispatch_data.items():
        for header in headers:
            header_location = header.get("header_location")
            dominant_action = header.get("dominant_action")
            dominant_target = header.get("dominant_target")

            for display in header.get("inline_displays", []):
                text_offset = parse_offset(display.get("text_offset"))
                if text_offset is None:
                    continue
                lookup.setdefault((filename, text_offset), []).append(
                    {
                        "source": "inline",
                        "header_location": header_location,
                        "action": display.get("action") or dominant_action,
                        "target": display.get("target") or dominant_target,
                        "command": display.get("block_command") or display.get("command"),
                        "opcode_location": display.get("opcode_location"),
                    }
                )

            for candidate in header.get("candidate_targets", []):
                for display in candidate.get("displays", []):
                    text_offset = parse_offset(display.get("text_offset"))
                    if text_offset is None:
                        continue
                    lookup.setdefault((filename, text_offset), []).append(
                        {
                            "source": "candidate",
                            "header_location": header_location,
                            "action": display.get("action") or candidate.get("inferred_action") or dominant_action,
                            "target": display.get("target") or candidate.get("inferred_target") or dominant_target,
                            "command": display.get("block_command") or display.get("command"),
                            "opcode_location": display.get("opcode_location"),
                            "table_location": candidate.get("table_location"),
                            "target_location": candidate.get("target_location"),
                        }
                    )
    return lookup


def build_target_lookup(
    entries: list[dict[str, Any]],
    *,
    location_key: str,
    extra_keys: list[str],
) -> dict[tuple[str, int], list[dict[str, Any]]]:
    lookup: dict[tuple[str, int], list[dict[str, Any]]] = {}
    for entry in entries:
        for match in entry.get("row_matches", []):
            filename = match.get("file") or entry.get("file")
            if not filename:
                continue
            offset = parse_offset(match.get("offset"))
            if offset is None:
                continue
            payload = {
                "location": entry.get(location_key),
                "command": match.get("block_command") or match.get("command"),
                "destination_label": entry.get("destination_label"),
            }
            for key in extra_keys:
                payload[key] = entry.get(key)
            if "preceding_text" in entry:
                payload["preceding_text"] = entry["preceding_text"]
            lookup.setdefault((filename, offset), []).append(payload)
    return lookup


def build_flag_lookup(flags_data: dict[str, dict[str, Any]]) -> dict[tuple[str, int], list[dict[str, Any]]]:
    lookup: dict[tuple[str, int], list[dict[str, Any]]] = {}
    for filename, payload in flags_data.items():
        for summary in payload.get("summary", []):
            for target in summary.get("targets", []):
                offset = parse_offset(target.get("offset"))
                if offset is None:
                    continue
                lookup.setdefault((filename, offset), []).append(
                    {
                        "kind": summary.get("kind"),
                        "arg1": summary.get("arg1"),
                        "arg2": summary.get("arg2"),
                        "count": summary.get("count"),
                        "locations": summary.get("locations", []),
                        "command": target.get("command"),
                    }
                )
    return lookup


def build_model(
    rows_by_file: dict[str, list[dict[str, Any]]],
    dispatch_lookup: dict[tuple[str, int], list[dict[str, Any]]],
    handoff_lookup: dict[tuple[str, int], list[dict[str, Any]]],
    transition_lookup: dict[tuple[str, int], list[dict[str, Any]]],
    flag_lookup: dict[tuple[str, int], list[dict[str, Any]]],
    only_context_rows: bool,
) -> dict[str, list[dict[str, Any]]]:
    model: dict[str, list[dict[str, Any]]] = {}

    for filename, rows in rows_by_file.items():
        output_rows: list[dict[str, Any]] = []
        for row in rows:
            key = (filename, row["offset"])
            dispatch_contexts = dispatch_lookup.get(key, [])
            handoff_contexts = handoff_lookup.get(key, [])
            transition_contexts = transition_lookup.get(key, [])
            flag_contexts = flag_lookup.get(key, [])

            if only_context_rows and not (
                row["display_type"]
                or row["dispatch_detail"]
                or row["handoff_detail"]
                or dispatch_contexts
                or handoff_contexts
                or transition_contexts
                or flag_contexts
            ):
                continue

            output_rows.append(
                {
                    **row,
                    "dispatch_contexts": dispatch_contexts,
                    "handoff_contexts": handoff_contexts,
                    "transition_contexts": transition_contexts,
                    "flag_contexts": flag_contexts,
                }
            )
        model[filename] = output_rows

    return model


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge workbook annotations and JSON analyzer outputs into one trigger model.")
    parser.add_argument("--dump", type=Path, default=Path(DUMP_XLS_PATH), help="Workbook to read (default: %(default)s)")
    parser.add_argument("--dispatch-json", type=Path, required=True, help="Path to dispatch_all.json")
    parser.add_argument("--flags-json", type=Path, required=True, help="Path to flags_all.json")
    parser.add_argument("--handoffs-json", type=Path, required=True, help="Path to handoffs_all.json")
    parser.add_argument("--transitions-json", type=Path, required=True, help="Path to transitions_all.json")
    parser.add_argument("--only-context-rows", action="store_true", help="Emit only rows that have trigger/context metadata")
    parser.add_argument("--output", type=Path, help="Write JSON to this path instead of stdout")
    args = parser.parse_args()

    rows_by_file = load_workbook_rows(args.dump)
    dispatch_lookup = build_dispatch_lookup(load_json(args.dispatch_json))
    handoff_lookup = build_target_lookup(load_json(args.handoffs_json), location_key="location", extra_keys=["arg1"])
    transition_lookup = build_target_lookup(load_json(args.transitions_json), location_key="location", extra_keys=["arg1"])
    flag_lookup = build_flag_lookup(load_json(args.flags_json))

    model = build_model(
        rows_by_file,
        dispatch_lookup,
        handoff_lookup,
        transition_lookup,
        flag_lookup,
        args.only_context_rows,
    )

    rendered = json.dumps(model, indent=2, ensure_ascii=False)
    if args.output:
        args.output.write_text(rendered, encoding="utf-8")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
