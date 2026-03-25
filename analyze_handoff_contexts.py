from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from rominfo import DUMP_XLS_PATH, ORIGINAL_ROM_DIR

SAVE_FORMAT_PATH = Path(__file__).resolve().parent / "docs" / "save_format.txt"


def parse_offset(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    return int(text, 16)


def hex5(value: int | None) -> str | None:
    if value is None:
        return None
    return f"0x{value:05x}"


def load_map_names(path: Path) -> dict[int, str]:
    if not path.exists():
        return {}

    result: dict[int, str] = {}
    pattern = re.compile(r"^\s*\*\s*([0-9a-fA-F]{2})\s*:\s*(.+?)\s*$")
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        match = pattern.match(raw_line)
        if not match:
            continue
        result[int(match.group(1), 16)] = match.group(2)
    return result


def load_dump_rows(path: Path) -> dict[tuple[str, int], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows_by_file_offset: dict[tuple[str, int], dict[str, Any]] = {}

    for sheet_name in workbook.sheetnames:
        if not sheet_name.endswith(".MSD"):
            continue
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(rows)]
        except StopIteration:
            continue
        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        if "Offset" not in name_to_index:
            continue

        current_command = ""
        for row in rows:
            offset = parse_offset(row[name_to_index["Offset"]])
            if offset is None:
                continue
            command = row[name_to_index["Command"]] if "Command" in name_to_index else None
            if command:
                current_command = str(command).strip()
            rows_by_file_offset[(sheet_name, offset)] = {
                "command": str(command).strip() if command else "",
                "block_command": current_command,
                "english": row[name_to_index["English"]] if "English" in name_to_index else None,
            }

    workbook.close()
    return rows_by_file_offset


def build_row_records(rows_by_file_offset: dict[tuple[str, int], dict[str, Any]], filename: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for (sheet_name, offset), row in rows_by_file_offset.items():
        if sheet_name != filename:
            continue
        records.append(
            {
                "file": sheet_name,
                "offset": offset,
                "command": row["command"],
                "block_command": row["block_command"],
                "english": row["english"],
            }
        )
    records.sort(key=lambda item: item["offset"])
    return records


def build_offset_to_rows(rows_by_file_offset: dict[tuple[str, int], dict[str, Any]]) -> dict[int, list[dict[str, Any]]]:
    result: dict[int, list[dict[str, Any]]] = {}
    for (sheet_name, offset), row in rows_by_file_offset.items():
        result.setdefault(offset, []).append(
            {
                "file": sheet_name,
                "offset": offset,
                "command": row["command"],
                "block_command": row["block_command"],
                "english": row["english"],
            }
        )
    return result


def infer_handoff_type(arg1: int, destination_name: str | None, command_label: str = "") -> str:
    destination_text = (destination_name or "").lower()
    command_text = command_label.lower()

    if "battle" in destination_text:
        return "battle handoff"
    if "scene" in destination_text:
        if "end of scene" in command_text:
            return "possible scene-end handoff"
        return "battle/scene-style handoff"
    if arg1 == 0x03:
        return "battle/scene-style handoff"
    return "transition handoff"


def format_destination_label(arg1: int, destination_name: str | None) -> str:
    if destination_name:
        return f'0x{arg1:02x} "{destination_name}"'
    return f"0x{arg1:02x}"


def summarize_handoff_details(handoff_contexts: list[dict[str, Any]]) -> list[str]:
    details: list[str] = []
    for context in handoff_contexts:
        opcode_location = context.get("location")
        opcode_text = (
            f"03 0x{context['arg1']:02x} 00 01 via {hex5(opcode_location)}"
            if isinstance(opcode_location, int)
            else f"03 0x{context['arg1']:02x} 00 01"
        )
        destination_label = str(context.get("destination_label") or "").strip()
        if destination_label:
            opcode_text += f" -> {destination_label}"
        details.append(opcode_text)
    return details


def find_preceding_text_opcode(
    data: bytes,
    handoff_location: int,
    offset_to_rows: dict[int, list[dict[str, Any]]],
    max_distance: int = 8,
) -> dict[str, int] | None:
    candidates: list[dict[str, int]] = []
    for start in range(handoff_location - 1, max(-1, handoff_location - max_distance - 1), -1):
        if start < 0 or start + 3 >= len(data):
            continue
        if data[start] == 0x02 and data[start + 1] != 0xFF and start + 4 <= handoff_location:
            candidates.append(
                {
                    "opcode_location": start,
                    "text_offset": data[start + 1] | (data[start + 2] << 8),
                    "arg": data[start + 3],
                    "delta": handoff_location - start,
                }
            )

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: (
            0 if item["text_offset"] in offset_to_rows else 1,
            item["delta"],
        )
    )
    return candidates[0]


def scan_handoffs(
    data: bytes,
    offset_to_rows: dict[int, list[dict[str, Any]]],
    map_names: dict[int, str],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for location in range(len(data) - 4):
        if data[location] != 0x03 or data[location + 2] != 0x00 or data[location + 3] != 0x01:
            continue

        preceding_text = find_preceding_text_opcode(data, location, offset_to_rows)
        row_matches = offset_to_rows.get(preceding_text["text_offset"], []) if preceding_text else []
        results.append(
            {
                "location": location,
                "arg1": data[location + 1],
                "destination_name": map_names.get(data[location + 1]),
                "destination_label": format_destination_label(data[location + 1], map_names.get(data[location + 1])),
                "preceding_text": preceding_text,
                "row_matches": row_matches,
                "window": data[max(0, location - 16) : min(len(data), location + 24)].hex(" "),
            }
        )
    return results


def scan_handoff_context_lookup(
    filename: str,
    data: bytes,
    row_records: list[dict[str, Any]],
    map_names: dict[int, str] | None = None,
) -> dict[int, list[dict[str, Any]]]:
    local_map_names = map_names if map_names is not None else load_map_names(SAVE_FORMAT_PATH)
    offset_to_rows = {
        row["offset"]: [
            {
                "file": row["file"],
                "offset": row["offset"],
                "command": row["command"],
                "block_command": row["block_command"],
                "english": row["english"],
            }
        ]
        for row in row_records
    }
    handoffs = scan_handoffs(data, offset_to_rows, local_map_names)
    lookup: dict[int, list[dict[str, Any]]] = {}
    for handoff in handoffs:
        preceding_text = handoff.get("preceding_text")
        if preceding_text is None:
            continue
        offset = preceding_text["text_offset"]
        if offset not in offset_to_rows:
            continue
        for row_match in handoff["row_matches"]:
            if row_match["file"] != filename:
                continue
            context = {
                "location": handoff["location"],
                "arg1": handoff["arg1"],
                "destination_name": handoff["destination_name"],
                "destination_label": handoff["destination_label"],
                "handoff_type": infer_handoff_type(
                    handoff["arg1"],
                    handoff["destination_name"],
                    row_match.get("block_command") or row_match.get("command") or "",
                ),
                "command": row_match.get("command") or "",
                "block_command": row_match.get("block_command") or "",
                "text_offset": offset,
            }
            lookup.setdefault(offset, []).append(context)
    return lookup


def human_report(results: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    lines.append(f"Total 03 xx 00 01 occurrences: {len(results)}")
    lines.append("")

    arg_counts = Counter(item["arg1"] for item in results)
    lines.append("Arg1 distribution:")
    for arg1, count in sorted(arg_counts.items()):
        lines.append(f"  0x{arg1:02x}: {count}")
    lines.append("")

    lines.append("Occurrences:")
    for item in results:
        lines.append(f"{hex5(item['location'])}: 03 0x{item['arg1']:02x} 00 01")
        lines.append(f"  Destination: {item['destination_label']}")
        preceding_text = item["preceding_text"]
        if preceding_text is not None:
            lines.append(
                f"  Preceding text: {hex5(preceding_text['opcode_location'])} -> "
                f"{hex5(preceding_text['text_offset'])} arg={preceding_text['arg']} "
                f"(delta={preceding_text['delta']})"
            )
        else:
            lines.append("  Preceding text: none within scan window")

        if item["row_matches"]:
            for match in item["row_matches"]:
                label = match["block_command"] or match["command"] or "(no command)"
                lines.append(f"  Workbook: {match['file']} {hex5(match['offset'])} -> {label}")
                if match["english"]:
                    lines.append(f"    English: {match['english']}")
        else:
            lines.append("  Workbook: no resolved row")
        lines.append(f"  Raw window: {item['window']}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def json_ready(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    payload = []
    for item in results:
        entry = {
            "location": hex5(item["location"]),
            "arg1": f"0x{item['arg1']:02x}",
            "destination_name": item["destination_name"],
            "destination_label": item["destination_label"],
            "row_matches": [
                {
                    "file": match["file"],
                    "offset": hex5(match["offset"]),
                    "command": match["command"],
                    "block_command": match["block_command"],
                    "english": match["english"],
                }
                for match in item["row_matches"]
            ],
            "window": item["window"],
        }
        if item["preceding_text"] is not None:
            entry["preceding_text"] = {
                "opcode_location": hex5(item["preceding_text"]["opcode_location"]),
                "text_offset": hex5(item["preceding_text"]["text_offset"]),
                "arg": item["preceding_text"]["arg"],
                "delta": item["preceding_text"]["delta"],
            }
        payload.append(entry)
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze the 03 xx 00 01 handoff/event-transition opcode family.")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of human-readable output")
    args = parser.parse_args()

    rows_by_file_offset = load_dump_rows(Path(DUMP_XLS_PATH))
    offset_to_rows = build_offset_to_rows(rows_by_file_offset)
    map_names = load_map_names(SAVE_FORMAT_PATH)
    pos_exe = (Path(ORIGINAL_ROM_DIR) / "POS.EXE").read_bytes()
    results = scan_handoffs(pos_exe, offset_to_rows, map_names)

    if args.json:
        print(json.dumps(json_ready(results), indent=2))
    else:
        print(human_report(results))


if __name__ == "__main__":
    main()
