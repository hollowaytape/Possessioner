from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rominfo import DUMP_XLS_PATH, MSD_POINTER_RANGES, ORIGINAL_ROM_DIR


DISPLAY_CALL = b"\x9a\x46\x22\xf0\x05"


def parse_offset(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    return int(text, 16)


def hex5(value: int) -> str:
    return f"0x{value:05x}"


def load_dump_rows(path: Path, file_filter: set[str]) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name not in file_filter:
            continue
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [cell for cell in next(rows)]
        except StopIteration:
            continue
        name_to_index = {name: idx for idx, name in enumerate(header) if name is not None}
        sheet_rows: list[dict[str, Any]] = []
        current_command = ""
        for row in rows:
            offset = parse_offset(row[name_to_index["Offset"]]) if "Offset" in name_to_index else None
            if offset is None:
                continue
            command = row[name_to_index["Command"]] if "Command" in name_to_index else None
            if command:
                current_command = str(command)
            sheet_rows.append(
                {
                    "offset": offset,
                    "command": str(command).strip() if command else "",
                    "block_command": current_command,
                    "ctrl_codes": row[name_to_index["Ctrl Codes"]] if "Ctrl Codes" in name_to_index else None,
                    "japanese": row[name_to_index["Japanese"]] if "Japanese" in name_to_index else None,
                    "english": row[name_to_index["English"]] if "English" in name_to_index else None,
                }
            )
        result[sheet_name] = sheet_rows

    workbook.close()
    return result


def build_offset_lookup(sheet_rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    return {row["offset"]: row for row in sheet_rows}


def find_next_display(
    data: bytes,
    start: int,
    stop: int,
) -> dict[str, Any] | None:
    cursor = start
    limit = min(stop, len(data) - 1)
    while cursor <= limit:
        if cursor + 3 < len(data) and data[cursor] == 0x02 and data[cursor + 1] != 0xFF:
            text_offset = data[cursor + 1] | (data[cursor + 2] << 8)
            return {
                "kind": "text",
                "opcode_location": cursor,
                "text_offset": text_offset,
                "arg": data[cursor + 3],
            }
        if (
            cursor + 10 < len(data)
            and data[cursor] == 0xBE
            and data[cursor + 3] == 0xB9
            and data[cursor + 6 : cursor + 11] == DISPLAY_CALL
        ):
            text_offset = data[cursor + 1] | (data[cursor + 2] << 8)
            row_count = data[cursor + 4] | (data[cursor + 5] << 8)
            return {
                "kind": "direct_range",
                "opcode_location": cursor,
                "text_offset": text_offset,
                "row_count": row_count,
            }
        cursor += 1
    return None


def scan_file_flags(
    filename: str,
    pos_exe: bytes,
    sheet_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    offsets = build_offset_lookup(sheet_rows)
    operations: list[dict[str, Any]] = []
    summary: dict[tuple[str, int, int], dict[str, Any]] = {}

    for range_start, range_stop in MSD_POINTER_RANGES[filename]:
        cursor = range_start
        while cursor <= min(range_stop, len(pos_exe) - 6):
            op_kind = None
            op_length = 0
            arg1 = None
            arg2 = None
            extra = None

            if pos_exe[cursor : cursor + 3] == b"\x02\xff\x01":
                op_kind = "set_flag"
                op_length = 5
                arg1 = pos_exe[cursor + 3]
                arg2 = pos_exe[cursor + 4]
            elif pos_exe[cursor : cursor + 3] == b"\x02\xff\x02":
                op_kind = "check_flag"
                op_length = 5
                arg1 = pos_exe[cursor + 3]
                arg2 = pos_exe[cursor + 4]
            elif pos_exe[cursor : cursor + 3] == b"\x03\xff\x01":
                op_kind = "clear_flag"
                op_length = 6
                arg1 = pos_exe[cursor + 3]
                arg2 = pos_exe[cursor + 4]
                extra = pos_exe[cursor + 5]

            if op_kind is None:
                cursor += 1
                continue

            display = find_next_display(pos_exe, cursor + op_length, min(cursor + 48, range_stop))
            target_row = offsets.get(display["text_offset"]) if display and "text_offset" in display else None
            op = {
                "file": filename,
                "location": cursor,
                "kind": op_kind,
                "arg1": arg1,
                "arg2": arg2,
                "extra": extra,
                "display": display,
                "target_command": target_row["command"] if target_row else "",
                "target_block_command": target_row["block_command"] if target_row else "",
                "target_offset": target_row["offset"] if target_row else None,
            }
            operations.append(op)

            key = (op_kind, arg1, arg2)
            if key not in summary:
                summary[key] = {
                    "kind": op_kind,
                    "arg1": arg1,
                    "arg2": arg2,
                    "count": 0,
                    "locations": [],
                    "targets": [],
                }
            summary[key]["count"] += 1
            summary[key]["locations"].append(cursor)
            target_label = target_row["block_command"] if target_row and target_row["block_command"] else ""
            if target_label:
                summary[key]["targets"].append(
                    {
                        "offset": target_row["offset"],
                        "command": target_label,
                    }
                )

            cursor += op_length

    deduped_summary = []
    for item in summary.values():
        seen_targets = set()
        targets = []
        for target in item["targets"]:
            key = (target["offset"], target["command"])
            if key in seen_targets:
                continue
            seen_targets.add(key)
            targets.append(target)
        deduped_summary.append(
            {
                "kind": item["kind"],
                "arg1": item["arg1"],
                "arg2": item["arg2"],
                "count": item["count"],
                "locations": [hex5(loc) for loc in item["locations"]],
                "targets": [
                    {"offset": hex5(target["offset"]), "command": target["command"]}
                    for target in targets
                ],
            }
        )

    deduped_summary.sort(key=lambda row: (row["kind"], row["arg1"], row["arg2"]))
    return {"operations": operations, "summary": deduped_summary}


def human_report(file_results: dict[str, dict[str, Any]]) -> str:
    lines: list[str] = []
    for filename, result in file_results.items():
        lines.append(f"{filename}")
        lines.append(f"  Operations: {len(result['operations'])}")
        counts = defaultdict(int)
        for op in result["operations"]:
            counts[op["kind"]] += 1
        lines.append(
            "  Counts: "
            + ", ".join(f"{kind}={counts.get(kind, 0)}" for kind in ("set_flag", "check_flag", "clear_flag"))
        )
        lines.append("  Distinct flag tuples:")
        for item in result["summary"]:
            targets = ", ".join(f"{t['command']} @ {t['offset']}" for t in item["targets"][:5]) or "(no resolved target)"
            lines.append(
                f"    {item['kind']} arg1=0x{item['arg1']:02x} arg2=0x{item['arg2']:02x} "
                f"x{item['count']} -> {targets}"
            )
        lines.append("  Raw operations:")
        for op in result["operations"]:
            display = op["display"]
            if display is None:
                display_text = "no nearby display op"
            elif display["kind"] == "text":
                display_text = f"text {hex5(display['text_offset'])} arg={display['arg']}"
            else:
                display_text = f"direct range {hex5(display['text_offset'])} rows={display['row_count']}"
            label = op["target_block_command"] or "(no block command)"
            extra = f" extra=0x{op['extra']:02x}" if op["extra"] is not None else ""
            lines.append(
                f"    {hex5(op['location'])}: {op['kind']} arg1=0x{op['arg1']:02x} arg2=0x{op['arg2']:02x}{extra} "
                f"-> {display_text} -> {label}"
            )
        lines.append("")
    return "\n".join(lines)


def json_ready(results: dict[str, dict[str, Any]]) -> dict[str, Any]:
    output = {}
    for filename, result in results.items():
        operations = []
        for op in result["operations"]:
            display = op["display"]
            op_json = {
                "location": hex5(op["location"]),
                "kind": op["kind"],
                "arg1": f"0x{op['arg1']:02x}",
                "arg2": f"0x{op['arg2']:02x}",
                "target_command": op["target_command"],
                "target_block_command": op["target_block_command"],
                "target_offset": hex5(op["target_offset"]) if op["target_offset"] is not None else None,
            }
            if op["extra"] is not None:
                op_json["extra"] = f"0x{op['extra']:02x}"
            if display is not None:
                op_json["display"] = {
                    **display,
                    "opcode_location": hex5(display["opcode_location"]),
                    "text_offset": hex5(display["text_offset"]),
                }
            operations.append(op_json)
        output[filename] = {"summary": result["summary"], "operations": operations}
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze flag operations in POS.EXE for selected MSD files.")
    parser.add_argument("--files", nargs="+", default=["POS1.MSD", "YUMI.MSD"])
    parser.add_argument("--json", action="store_true", help="Output JSON instead of human-readable text")
    args = parser.parse_args()

    file_filter = set(args.files)
    dump_rows = load_dump_rows(Path(DUMP_XLS_PATH), file_filter)
    pos_exe = (Path(ORIGINAL_ROM_DIR) / "POS.EXE").read_bytes()

    results = {
        filename: scan_file_flags(filename, pos_exe, dump_rows.get(filename, []))
        for filename in args.files
    }

    if args.json:
        print(json.dumps(json_ready(results), indent=2))
    else:
        print(human_report(results))


if __name__ == "__main__":
    main()
