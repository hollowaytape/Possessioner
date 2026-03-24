import struct
from openpyxl import load_workbook

# Load MSD pointer ranges from rominfo
msd_pointer_ranges = {
    "POS1.MSD": [
        (0x01ec, 0x0209),
        (0x035f, 0x0384),
        (0x0500, 0x052f),
        (0x0749, 0x078c),
        (0x08ea, 0x0958),
        (0x0b18, 0x0c43),
        (0x0d14, 0x0f5f),
        (0x1009, 0x1184),
        (0x1290, 0x1453),
        (0x1554, 0x1722),
        (0x185a, 0x1963),
    ],
    "YUMI.MSD": [
        (0x027a8, 0x027b4),
        (0x02803, 0x0281d),
        (0x0289d, 0x028c9),
        (0x02abc, 0x02ace),
    ],
}

# For each file, find which flags appear in pointer contexts
wb = load_workbook("D:\\Code\\roms\\Possessioner\\PSSR_dump.xlsx")

with open("D:\\Code\\roms\\Possessioner\\original\\POS.EXE", "rb") as f:
    exe_data = f.read()

# Map pointer locations to MSD file/offsets from workbooks
file_pointer_map = {}  # {pointer_loc: [(msd_file, offset, command), ...]}

for msd_file in ["POS1.MSD", "YUMI.MSD"]:
    if msd_file not in wb.sheetnames:
        continue
    
    ws = wb[msd_file]
    
    # Check if there's a pointer column
    has_pointer = False
    pointer_col = None
    offset_col = None
    command_col = None
    
    for col_num in range(1, 50):
        header = ws.cell(1, col_num).value
        if header and "Pointer" in str(header):
            pointer_col = col_num
            has_pointer = True
        if header and "Offset" in str(header):
            offset_col = col_num
        if header and "Command" in str(header):
            command_col = col_num
    
    if not has_pointer:
        # Try to get pointers from a different sheet
        if "Pointer Dump" in msd_file.lower():
            continue

print("\\n=== Summary for Analysis ===")
print(f"POS1.MSD pointer ranges (in POS.EXE): {msd_pointer_ranges['POS1.MSD']}")
print(f"YUMI.MSD pointer ranges (in POS.EXE): {msd_pointer_ranges['YUMI.MSD']}")

# Now let's find all flag operations and see which MSD file they belong to
print("\n=== Flag Operations in POS.EXE ===\n")

set_flags_in_range = {"POS1.MSD": [], "YUMI.MSD": []}

for i in range(len(exe_data) - 4):
    if exe_data[i] == 0x02 and exe_data[i+1] == 0xff and exe_data[i+2] == 0x01:
        # SET_FLAG
        flag_id = exe_data[i+3]
        flag_bit = exe_data[i+4]
        
        # Check which MSD file range this belongs to
        for msd_file, ranges in msd_pointer_ranges.items():
            for start, end in ranges:
                if start <= i <= end:
                    set_flags_in_range[msd_file].append((i, flag_id, flag_bit))
                    break

print("SET_FLAG operations by file:")
for msd_file in ["POS1.MSD", "YUMI.MSD"]:
    print(f"\n{msd_file}:")
    if set_flags_in_range[msd_file]:
        for ptr_loc, flag_id, flag_bit in set_flags_in_range[msd_file]:
            print(f"  0x{ptr_loc:05x}: flag_id=0x{flag_id:02x}, bit={flag_bit}")
    else:
        print(f"  (None found in pointer ranges)")
